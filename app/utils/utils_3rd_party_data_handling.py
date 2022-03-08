from copy import deepcopy

from flask import flash, Response
from kafka import KafkaProducer
from kafka.errors import KafkaError
from kafka.oauth import AbstractTokenProvider
from sqlalchemy.exc import SQLAlchemyError

from app.globals import GLOBAL_IP
from app.models import *
# from app.csv_to_json_converter_util import *
from sqlalchemy import exists, func
from datetime import datetime
import openpyxl
import json
import os
import requests
import stix2
import stix2validator
import app.utils.stix2_custom as stix2_custom

# region Insert all CAPEC records from Excel
# from app.utils.utils_communication import send_alert_info_update_needed
# from app.producer import SendKafkaReport
from app.utils.utils_risk_assessment import start_risk_assessment_alert, risk_assessment_save_report


# def start_risk_assessment_alert ():
#     pass

# ################ SHOULD BE REMOVED AND USE SendKafkaReport from producer #################
path_to_kafka_cert = os.path.join(os.path.abspath(os.getcwd()), 'app', 'auth_files', 'for_clients.crt')

SM_IP = os.environ.get('SM_IP') if os.environ.get(
    'SM_IP') else "http://sphinx-toolkit.intracom-telecom.com/SMPlatform/manager/rst"
KAFKA_USERNAME = os.environ.get('KAFKA_USERNAME') if os.environ.get('KAFKA_USERNAME') else "kafkauser"
KAFKA_PASSWORD = os.environ.get('KAFKA_PASSWORD') if os.environ.get('KAFKA_PASSWORD') else "kafkauser123"
OAUTH_CLIENT_ID = os.environ.get('OAUTH_CLIENT_ID') if os.environ.get('OAUTH_CLIENT_ID') else "SIEM"
OAUTH_TOKEN_ENDPOINT_URI = os.environ.get('OAUTH_TOKEN_ENDPOINT_URI') if os.environ.get(
    'OAUTH_TOKEN_ENDPOINT_URI') else "http://sphinx-toolkit.intracom-telecom.com/SMPlatform/manager/rst/getKafkaToken"
BOOTSTRAP_SERVERS = os.environ.get('BOOTSTRAP_SERVERS') if os.environ.get(
    'BOOTSTRAP_SERVERS') else "'bootstrap.146.124.106.181.nip.io:443"
KAFKA_CERT = os.environ.get('KAFKA_CERT')  # FULL PATH OF THE CERTIFICATE LOCATION

class TokenProvider(AbstractTokenProvider):

    def __init__(self):
        self.kafka_ticket = json.loads(requests.post(SM_IP + '/KafkaAuthentication', data={'username': KAFKA_USERNAME,
                                                                                           'password': KAFKA_PASSWORD}).text)[
            'data']

    def token(self):
        kafka_token = \
        json.loads(requests.get(OAUTH_TOKEN_ENDPOINT_URI, auth=(OAUTH_CLIENT_ID, self.kafka_ticket)).text)[
            'access_token']

        return kafka_token

def SendKafkaReport(report, topic_to_write):
    # return ;
    # KAFKA CLIENT PRODUCER
    print("Initialising Kafka Producer")
    producer = KafkaProducer(bootstrap_servers=BOOTSTRAP_SERVERS,
                             security_protocol='SASL_SSL',
                             sasl_mechanism='OAUTHBEARER',
                             sasl_oauth_token_provider=TokenProvider(),
                             ssl_cafile=path_to_kafka_cert,
                             value_serializer=lambda value: value.encode(),
                             api_version=(2, 5, 0))
    # print(BOOTSTRAP_SERVERS)
    # print(os.environ.get('BOOTSTRAP_SERVERS'))
    # # producer = KafkaProducer(bootstrap_servers=BOOTSTRAP_SERVERS,
    # #                         security_protocol='SASL_SSL',
    # #                         sasl_mechanism='OAUTHBEARER',
    # #                         sasl_oauth_token_provider=TokenProvider(),
    # #                         ssl_cafile= path_to_kafka_cert,
    # #                         value_serializer=lambda value: value.encode())
    #
    #
    print("Trying to send with Kafka Producer")
    try:
        producer.send(topic_to_write, json.dumps(report))
    except KafkaError:
        print("Kafka producing sending data encountered an error")

    result = producer.flush()
    print(result, flush=True)
    producer.close()

# ############################## UNITL HERE ######################################
def send_risk_report(report_id, asset_id, threat_id):
    try:
        this_asset = RepoAsset.query.filter_by(id=asset_id).first()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)

    try:
        this_risk_assessment_report = RepoRiskAssessmentReports.query.filter_by(id=report_id).first()
    except SQLAlchemyError:
        return "SQLAlchemyError"

    try:
        this_threat = RepoThreat.query.filter_by(id=threat_id).first()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)

    try:
        this_threat_asset_exposure = RepoAssetRepoThreatRelationship.query.filter_by(repo_asset_id=asset_id,
                                                                                     repo_threat_id=threat_id).first()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)

    report_to_send = {}
    report_to_send["report_info"] = {
        "date_time": this_risk_assessment_report.date_time.strftime("%m/%d/%Y, %H:%M:%S"),
        "type": this_risk_assessment_report.type
    }

    try:
        these_vulnerabilities = VulnerabilityReportVulnerabilitiesLink.query.filter_by(
                        asset_id=asset_id).all()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)



    report_to_send["assset"] = {
        "id": this_asset.id,
        "name": this_asset.name,
        # "asset_reputation": 0,  # placeholder
        "ip": this_asset.ip,
        "mac": this_asset.mac_address,
        "last_touched": this_asset.last_touch_date,
        # "vulnerabilities" : [
        #     {
        #         "cve_id": "CVE-2020-0645",
        #         "controls": []
        #     },
        #     {
        #         "cve_id": "CVE-2020-0774",
        #         "controls": [ {"description":"Updated software",
        #                        "effectiveness" : "high"
        #                        }]
        #     }
        # ],

        # "type": this_asset.type, #aSSETS DONT HAVE TYPE SHOULD BE ADDED
        # "related_services": SHould this be added?
    }

    vulnerabilities_to_add = []
    for vulnerability in these_vulnerabilities:
        temp_to_add = {"cve_id": vulnerability.cve.CVEId, "controls": []}
        try:
            these_controls = RepoControl.query.filter_by( vulnerability_id=vulnerability.id).all()
        except SQLAlchemyError:
            return Response("SQLAlchemyError", 500)

        for control in these_controls:
            temp_effectiveness = ""
            if control.effectiveness:
                temp_effectiveness = str(control.effectiveness)

            temp_to_add["controls"].append({"description": control.name,
                               "effectiveness": temp_effectiveness
                               })

        vulnerabilities_to_add.append(temp_to_add)

    report_to_send["vulnerabilities"] = vulnerabilities_to_add

    report_to_send["threat"] = {
        "name": this_threat.name,
        "capec_info": {
            "capec_id": "",
            "name": "",
            "abstraction": "",
            "likelihood": "",
            "typical_severity": ""
        },
        "threat_asset_info": {
            "skill_level": this_threat_asset_exposure.risk_skill_level,
            "motive": this_threat_asset_exposure.risk_motive,
            "source": this_threat_asset_exposure.risk_source,
            "actor": this_threat_asset_exposure.risk_actor,
            "opportunity": this_threat_asset_exposure.risk_opportunity,
        }
    }

    exposure_inference_values = this_risk_assessment_report.exposure_inference.split("|")
    objectives_inference_values = this_risk_assessment_report.objectives_inference.split("|")
    utility_inference_values = this_risk_assessment_report.utilities_inference.split("|")
    alerts_triggered = this_risk_assessment_report.alerts_triggered.split("|")
    # utility_inference_values = this_risk_assessment_report.responses_inference.split("|") # TODO Change to correct field after the model is fixed itself
    static_info_to_add = {}
    # Load static info to the report
    # exposure_set = []
    # materialisations_set = []
    # responses_set = []
    # consequences_set = []
    # services_set = []
    # impacts_set = []
    # objectives_set = []
    if this_risk_assessment_report.exposure_set:
        exposure_to_add = {}
        exposure_set = this_risk_assessment_report.exposure_set.split("|")
        for it in range(0, len(exposure_set) - 1, 2):
            exposure_to_add[this_threat.name] = exposure_set[it + 1]

        static_info_to_add["exposure"] = exposure_to_add

    static_info_to_add = {
        "service_insurance_check" : "1",
        # "threat_occurance": "1",
        # "materialisation": "1",
        # "Unauthorised modifications of data": "1",
        # "Under maintenance": "0",
        # "Safety": "0",
        # "Integrity": "0"
    }
    if this_risk_assessment_report.materialisations_set:
        materialisation_to_add = {}
        materialisations_set = this_risk_assessment_report.materialisations_set.split("|")
        for it in range(0, len(materialisations_set) - 1, 2):
            try:
                this_materialisation = RepoMaterialisation.query.filter_by(id=materialisations_set[it]).first()
            except SQLAlchemyError:
                return Response("SQLAlchemyError", 500)

            materialisation_to_add[this_materialisation.name] = materialisations_set[it + 1]
        static_info_to_add["materialisations"] = materialisation_to_add


    alerts_to_add = []
    for alert in alerts_triggered:
        if alert == "":
            continue
        print("-------ALERT WITH ERROR")
        print(alert)
        print(type(alert))
        alerts_to_add.append(json.loads(alert))


    # Need to add the other static info
    report_to_send["risk"] = {
        "static_info": static_info_to_add,
        "exposure_threat": {
            "occurrence": str(exposure_inference_values[2])
        },
        "objectives": {
            "confidentiality": {
                "low": str(objectives_inference_values[1]),
                "medium": str(objectives_inference_values[2]),
                "high": str(objectives_inference_values[3])
            },
            "integrity": {
                "low": str(objectives_inference_values[5]),
                "medium": str(objectives_inference_values[6]),
                "high": str(objectives_inference_values[7])
            },
            "availability": {
                "low": str(objectives_inference_values[9]),
                "medium": str(objectives_inference_values[10]),
                "high": str(objectives_inference_values[11])
            },
            "monetary": {
                "low": str(objectives_inference_values[13]),
                "medium": str(objectives_inference_values[14]),
                "high": str(objectives_inference_values[15])
            },
            "safety": {
                "low": str(objectives_inference_values[17]),
                "medium": str(objectives_inference_values[18]),
                "high": str(objectives_inference_values[19])
            },
            #
            # },
            "utilities": {
                    "CIA" : { "optimal_scenario" : json.loads(utility_inference_values[0])["optimal_scenario"], "most_probable_scenarios": json.loads(utility_inference_values[1])["most_probable_scenarios"]},
                    "Evaluation" : { "optimal_scenario" : json.loads(utility_inference_values[2])["optimal_scenario"], "most_probable_scenarios": json.loads(utility_inference_values[3])["most_probable_scenarios"]},
                    # "Evaluation" : [json.loads(utility_inference_values[1]), json.loads(utility_inference_values[2])]
            },
            "alerts" : alerts_to_add,
            # "utilities": {
            #     "CIA": {
            #         "most_probable_scenarios" : [
            #             {
            #                 "confidentiality" : "medium",
            #                 "integrity" : "medium",
            #                 "availability" : "low",
            #                 "probability" : "0.2891"
            #
            #             },
            #             {
            #                 "confidentiality": "high",
            #                 "integrity": "high",
            #                 "availability": "medium",
            #                 "probability": "0.2654"
            #
            #             },
            #             {
            #                 "confidentiality": "medium",
            #                 "integrity": "medium",
            #                 "availability": "medium",
            #                 "probability": "0.1266"
            #
            #             },
            #         ],
            #         "optimal_scenario":{
            #             "confidentiality": "low",
            #             "integrity": "low",
            #             "availability": "low",
            #             "probability": "0.0225"
            #         }
            #     },
            #     "Evaluation":{
            #         "most_probable_scenarios" : [
            #             {
            #                 "monetary" : "low",
            #                 "safety" : "low",
            #                 "probability" : "0.6275"
            #             },
            #             {
            #                 "monetary" : "medium",
            #                 "safety" : "medium",
            #                 "probability": "0.1573"
            #
            #             },
            #             {
            #                 "monetary" : "low",
            #                 "safety" : "medium",
            #                 "probability": "0.0853"
            #             },
            #         ],
            #         "optimal_scenario":{
            #             "monetary" : "low",
            #                 "safety" : "low",
            #                 "probability" : "0.6275"
            #         }
            #     },
            # },
            # "alerts": {
            #     "objectives": {
            #         "confidentiality": {
            #             "level" : "high",
            #             "threshold" : "0.4"
            #         }
            #     }
            # }
        }
    }

    print("----- THE REPORT IS -----")
    print(report_to_send)
    print(json.dumps(report_to_send))
    with open('example_output.json', 'w', encoding='utf-8') as f:
        json.dump(report_to_send, f, ensure_ascii=False, indent=4)
    report_to_send = json.dumps(report_to_send)

    # print(report_to_send)
    # SendKafkaReport(report_to_send, "rcra-report-topic")


def send_alert_new_asset(asset_id):
    now = datetime.now()

    try:
        asset_obj = RepoAsset.query.filter_by(id=asset_id).first()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)

    print("1")
    try:
        asset_vulnerabilities_count = VulnerabilityReportVulnerabilitiesLink.query.join(RepoAsset).filter(
            RepoAsset.id == asset_obj.id).count()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)
    print("2")
    alert_to_send = {
        "alert_type": "new_asset_detected",
        "date_time": now.strftime("%m/%d/%Y, %H:%M:%S"),
        "asset": {
            "asset_ip": asset_obj.ip if asset_obj.ip else "",
            "asset_common_id": asset_obj.common_id if asset_obj.common_id else ""
        },
        "asset_url": GLOBAL_IP + "repo/assets/" + str(asset_obj.id) + "/"
    }

    print("Alerts is --------", alert_to_send, flush=True)
    with open('send_alert_new_asset'+ str(asset_id) +'.json', 'w', encoding='utf-8') as f:
        json.dump(alert_to_send, f, ensure_ascii=False, indent=4)

    # SendKafkaReport(alert_to_send, "rcra-report-topic")
    return alert_to_send

def send_alert_info_update_needed(asset_id=None, threat_id=None, threat_exposure_info=-1,
                                  threat_materialisation_info=-1, threat_impact_info=-1, objective_info=-1,
                                  utility_info=-1):
    now = datetime.now()

    try:
        asset_obj = RepoAsset.query.filter_by(id=asset_id).first()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)

    try:
        threat_obj = RepoThreat.query.filter_by(id=threat_id).first()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)

    pages_to_send = []
    # Add exposure info in alert
    if threat_exposure_info != -1:
        pages_to_send.append(
            GLOBAL_IP + "repo/risk/configuration/threat/exposure/" + str(threat_obj.id) + "/asset/" + str(asset_obj.id) + "/")

    # Add mat and cons info in alert
    if threat_materialisation_info != -1:
        pages_to_send.append(
            GLOBAL_IP + "repo/risk/configuration/threat/" + str(threat_obj.id) + "/asset/" + str(asset_obj.id) + "/")

    # Add impact info in alert
    if threat_impact_info != -1:
        pages_to_send.append(
            GLOBAL_IP + "repo/risk/configuration/impact/1/threat" + str(threat_obj.id) + "/asset/" + str(asset_obj.id) + "/")
        pages_to_send.append(
            GLOBAL_IP + "repo/risk/configuration/impact/2/threat" + str(threat_obj.id) + "/asset/" + str(asset_obj.id) + "/")
        pages_to_send.append(
            GLOBAL_IP + "repo/risk/configuration/impact/3/threat" + str(threat_obj.id) + "/asset/" + str(asset_obj.id) + "/")
        pages_to_send.append(
            GLOBAL_IP + "repo/risk/configuration/impact/4/threat" + str(threat_obj.id) + "/asset/" + str(asset_obj.id) + "/")
        pages_to_send.append(
            GLOBAL_IP + "repo/risk/configuration/impact/5/threat" + str(threat_obj.id) + "/asset/" + str(asset_obj.id) + "/")

    # Add Objective info in alert
    if objective_info != -1:
        pages_to_send.append(GLOBAL_IP + "repo/risk/configuration/objective/1/")
        pages_to_send.append(GLOBAL_IP + "repo/risk/configuration/objective/2/")
        pages_to_send.append(GLOBAL_IP + "repo/risk/configuration/objective/3/")
        pages_to_send.append(GLOBAL_IP + "repo/risk/configuration/objective/4/")
        pages_to_send.append(GLOBAL_IP + "repo/risk/configuration/objective/5/")

    # Add Utility info in alert
    if utility_info != -1:
        pages_to_send.append(GLOBAL_IP + "/repo/risk/configuration/utility/1/")
        pages_to_send.append(GLOBAL_IP + "/repo/risk/configuration/utility/2/")

    alert_to_send = {
        "alert_type": "risk_assessment_info_update_needed",
        "date_time": now.strftime("%m/%d/%Y, %H:%M:%S"),
        "asset": {
            "asset_ip": asset_obj.ip if asset_obj.ip else "",
            "asset_common_id": asset_obj.common_id if asset_obj.common_id else "",
        },
        "threat": threat_obj.name,
        "pages_update_url": pages_to_send
    }

    print("Alerts is --------", alert_to_send, flush=True)
    with open('send_alert_info_update_needed' + str(asset_id) + '.json', 'w', encoding='utf-8') as f:
        json.dump(alert_to_send, f, ensure_ascii=False, indent=4)
    return alert_to_send


def security_event_risk_reports(report_id):
    now = datetime.now()

    try:
        report_obj = RepoRiskAssessmentReports.query.filter_by(id=report_id).first()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)

    try:
        asset_obj = RepoAsset.query.filter_by(id=report_obj.risk_assessment.repo_asset_id).first()
    except SQLAlchemyError:
        return Response("SQLAlchemyError", 500)

    alert_to_send = {
        "alert_type": "security_event_risk_reports",
        "alert_severity": 3,
        "date_time": now.strftime("%m/%d/%Y, %H:%M:%S"),
        "asset": {
            "asset_ip": asset_obj.ip if asset_obj.ip else "",
            "asset_common_id": asset_obj.common_id if asset_obj.common_id else "",
        },
        "risk_reports": [
            {
                "report_url": GLOBAL_IP + "repo/dashboard/risk/objectives/threat/"+ str(report_obj.risk_assessment.repo_threat_id) +"/asset/"+ str(report_obj.risk_assessment.repo_asset_id) +"/assessment/"+ str(report_id) + "/"
            }
        ]
    }

    print("Alerts is --------", alert_to_send, flush=True)
    with open('security_event_risk_reports' + str(report_id) + '.json', 'w', encoding='utf-8') as f:
        json.dump(alert_to_send, f, ensure_ascii=False, indent=4)




def CAPEC_excel_insertData(capecexcelpath):
    theFile = openpyxl.load_workbook(capecexcelpath)
    currentSheet = theFile.active
    for row in currentSheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # We need to check that the cell is not empty.
            if not db.session.query(
                    exists().where(CommonAttackPatternEnumerationClassification.capecId == row[0])).scalar():
                my_row = CommonAttackPatternEnumerationClassification(capecId=row[0], name=row[1], abstraction=row[2],
                                                                      status=row[3], description=row[4],
                                                                      alternateTerms=row[5], likelihoodOfAttack=row[6],
                                                                      typicalSeverity=row[7],
                                                                      relatedAttackpatterns=row[8],
                                                                      executionFlow=row[9], prerequisites=row[10],
                                                                      skillsRequired=row[11], resourcesRequired=row[12],
                                                                      indicators=row[13], consequences=row[14],
                                                                      mitigations=row[15], exampleInstances=row[16],
                                                                      relatedWeaknesses=row[17],
                                                                      taxonomyMappings=row[18], notes=row[19])
                db.session.add(my_row)
    db.session.commit()
    return 1

# endregion CAPEC

# region Insert all CWE records from Excel
def CWE_excel_insertData(cweexcelpath):
    theFile = openpyxl.load_workbook(cweexcelpath)
    currentSheet = theFile.active
    for row in currentSheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            row_cweID = str(row[0])
            if db.session.query(exists().where(CommonWeaknessEnumeration.CWEId == row_cweID)).scalar():
                my_row = db.session.query(CommonWeaknessEnumeration).filter_by(CWEId=row_cweID)[0]
            else:
                my_row = CommonWeaknessEnumeration(CWEId=row[0])
            my_row.name = row[1]
            my_row.weakness = row[2]
            my_row.abstraction = row[3]
            my_row.status = row[4]
            my_row.description = row[5]
            my_row.extendedDescription = row[6]
            my_row.relatedWeaknesses = row[7]
            my_row.weaknessOrdinalities = row[8]
            my_row.applicablePlatforms = row[9]
            my_row.backgroundDetails = row[10]
            my_row.alternateTerms = row[11]
            my_row.modesOfIntroduction = row[12]
            my_row.exploitationFactors = row[13]
            my_row.likelihoodOfExploit = row[14]
            my_row.commonConsequences = row[15]
            my_row.detectionMethods = row[16]
            my_row.potentialMitigations = row[17]
            my_row.observedExamples = row[18]
            my_row.functionalAreas = row[19]
            my_row.affectedResources = row[20]
            my_row.taxonomyMappings = row[21]
            my_row.relatedAttackPatterns = row[22]
            my_row.notes = row[23]
            db.session.add(my_row)
    db.session.commit()
    return 1

# endregion CWE

# region Insert all CVE records from Excel
def CVE_excel_insertData(cveexcelpath):
    theFile = openpyxl.load_workbook(cveexcelpath)
    currentSheet = theFile.active
    for row in currentSheet.iter_rows(min_row=2, values_only=True):
        # print(row[0])
        if row[0] is not None:
            if not db.session.query(
                    exists().where(CommonVulnerabilitiesAndExposures.CVEId == row[0])).scalar():
                my_row = CommonVulnerabilitiesAndExposures(CVEId=row[0], status=row[1])
                db.session.add(my_row)
    db.session.commit()
    return 1

# endregion CVE
# endregion Insert information from Excel files

# region Insert information from VAaaS Report
def v_report(fpath):
    with open(fpath, "r") as fp:
        obj = json.load(fp)
        print("VRERPORT")
        print(obj)
        print(type(obj))
        if obj["id"] is not None:
            reprow_reportId = obj["id"]
            if db.session.query(VulnerabilityReport.id).filter_by(reportId=reprow_reportId).first() is not None:
                my_json_report = db.session.query(VulnerabilityReport).filter_by(reportId=reprow_reportId).one()
            else:
                my_json_report = VulnerabilityReport(reportId=reprow_reportId)
            my_json_report.scan_start_time = obj["scan_start_time"] if obj["scan_start_time"] is not None else ""
            my_json_report.scan_end_time = obj["scan_end_time"] if obj["scan_end_time"] is not None else ""
            my_json_report.target_name = obj["target_name"] if obj["target_name"] is not None else ""
            my_json_report.source_component = 1
            db.session.add(my_json_report)
            try:
                db.session.commit()
                # flash('Vulnerability Report "{}" Added Succesfully'.format(my_json_report.reportId))
            except SQLAlchemyError as e:
                db.session.rollback()
                return -1
            # Get asset IP
            for item in obj['objects']:
                if item['type'] != "ipv4-addr":
                    continue
                else:
                    my_asset_IP = item["value"]
                    if db.session.query(RepoAsset.id).filter_by(ip=my_asset_IP).first() is None:
                        my_repo_asset = RepoAsset(ip=my_asset_IP)
                        db.session.add(my_repo_asset)
                        try:
                            db.session.commit()
                            # flash('Asset "{}" Added Succesfully'.format(my_repo_asset.ip))
                            #TODO: Send alert for the new Asset to the EndUser
                        except SQLAlchemyError as e:
                            db.session.rollback()
                            continue
                    else:
                        my_repo_asset = db.session.query(RepoAsset).filter_by(ip=my_asset_IP).first()

            # Get CVE from the result nodes of the report
            for item in obj['objects']:
                if item['type'] != "vulnerability":
                    continue
                else:
                    if item["cvss"] == "0.0":
                        continue
                    else:
                        for subitem in item['external_references']:
                            if subitem['source_name'] == "cve":
                                reprow_cveId = subitem['external_id']
                                if db.session.query(CommonVulnerabilitiesAndExposures.id).filter_by(CVEId=reprow_cveId).first() is None:
                                    my_cve = CommonVulnerabilitiesAndExposures(CVEId=reprow_cveId)
                                    db.session.add(my_cve)
                                else:
                                    my_cve = db.session.query(CommonVulnerabilitiesAndExposures).filter_by(CVEId=reprow_cveId).one()

                                if VulnerabilityReport.query.join(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter((VulnerabilityReportVulnerabilitiesLink.vreport_id == my_json_report.id) & (VulnerabilityReportVulnerabilitiesLink.cve_id == my_cve.id)).first() is None:
                                    my_link = VulnerabilityReportVulnerabilitiesLink(vreport_id=my_json_report.id, cve_id=my_cve.id)
                                    db.session.add(my_link)
                                else:
                                    my_link = VulnerabilityReport.query.join(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter((VulnerabilityReportVulnerabilitiesLink.vreport_id == my_json_report.id) & (VulnerabilityReportVulnerabilitiesLink.cve_id == my_cve.id)).first()
                                my_link.asset_id = my_repo_asset.id
                                my_link.VReport_assetID = obj['target_id'] if obj['target_id'] is not None else ""
                                my_link.VReport_assetIp = my_asset_IP if my_asset_IP is not None else ""
                                my_link.VReport_port = item['vulnerable_port'] if item['vulnerable_port'] is not None else ""
                                my_link.VReport_CVSS_score = item['cvss'] if item['cvss'] is not None else ""
                                my_link.comments = item['threat_level'] if item['threat_level'] is not None else ""
                                try:
                                    db.session.commit()
                                    # flash('Vulnerability "" Added Succesfully')
                                    # flash('Vulnerability "{}" Added Succesfully'.format(my_link.cve_id))
                                except SQLAlchemyError as e:
                                    db.session.rollback()
                                    continue
                                # update_cve_scores(reprow_cveId)
                                # TODO: It's not needed to call the update CVE and CWE functions
            return 1

def v_report_json(report_name, report_details):
    # obj = json.load(fp)
    obj = report_details
    print("VRERPORT")
    print(obj)
    print(type(obj))
    if obj["id"] is not None:
        reprow_reportId = obj["id"]
        if db.session.query(VulnerabilityReport.id).filter_by(reportId = reprow_reportId).first() is not None:
            my_json_report = db.session.query(VulnerabilityReport).filter_by(reportId=reprow_reportId).one()
        else:
            my_json_report = VulnerabilityReport(reportId=reprow_reportId)
        my_json_report.scan_start_time = obj["start"] if obj["start"] is not None else ""
        my_json_report.scan_end_time = obj["stop"] if obj["stop"] is not None else ""
        my_json_report.target_name = obj["task_name"] if obj["task_name"] is not None else ""
        my_json_report.assessment_date = obj["assessment_date"] if obj["assessment_date"] is not None else ""
        my_json_report.cvss_score = obj["cvss_score"] if obj["cvss_score"] is not None else ""
        my_json_report.total_services = obj["total_services"] if obj["total_services"] is not None else ""
        my_json_report.source_component = 1
        db.session.add(my_json_report)
        try:
            db.session.commit()
            # flash('Vulnerability Report "{}" Added Succesfully'.format(my_json_report.reportId))
        except SQLAlchemyError as e:
            db.session.rollback()
            return -1

        # Get asset IP
        my_asset_IP = None
        my_asset_MAC = None
        for item in obj['objects']:
            if item['type'] == "ipv4-addr":
                my_asset_IP = item["value"]
            elif item[''] != "":
                my_asset_MAC = item["value"]
            else:
                continue
        # TODO: Change to search with MAC address???
        if not db.session.query(RepoAsset.id).filter_by(ip = my_asset_IP).first() in None:
            my_repo_asset = RepoAsset(ip=my_asset_IP)
            my_repo_asset.mac_address = my_asset_MAC if my_asset_MAC is not None else ""
            db.session.add(my_repo_asset)
            try:
                db.session.commit()
                flash('Asset "{}" Added Succesfully'.format(my_repo_asset.ip))
                # TODO: Send alert for the new Asset to the EndUser
            except SQLAlchemyError as e:
                db.session.rollback()

            send_alert_new_asset(my_repo_asset.id)
        else:
            my_repo_asset = db.session.query(RepoAsset).filter_by(ip = my_asset_IP).first()
            my_repo_asset.mac_address = my_asset_MAC if my_asset_MAC is not None else ""

        # Get CVE from the result nodes of the report
        for item in obj['objects']:
            if item['type'] == "x-discovered-service":
                reprow_service_name = item['service_name']
                if db.session.query(RepoAssetService.id).filter_by(asset_id = my_repo_asset.ip, service_name = reprow_service_name).first() is None:
                    my_asset_service = RepoAssetService(asset_id=my_repo_asset.id, service_name=reprow_service_name)
                    db.session.add(my_asset_service)
                else:
                    my_asset_service = db.session.query(RepoAssetService).filter_by(asset_id=reprow_cveId, service_name=reprow_service_name).one()
                my_asset_service.vreport_id = my_json_report.id if my_json_report.id is not None else ""
                my_asset_service.port = item['port'] if item['port'] is not None else ""
                my_asset_service.protocol = item['protocol'] if item['protocol'] is not None else ""
                my_asset_service.state = item['state'] if item['state'] is not None else ""
                my_asset_service.service_product = item['service_product'] if item['service_product'] is not None else ""
                my_asset_service.service_product_version = item['service_product_version'] if item['service_product_version'] is not None else ""

                for service_attr, service_data in item['service_vulnerabilities'].items():
                    for vulnerability_item in service_data['null']:
                        if vulnerability_item['type'] == "cve":
                            reprow_cveId = vulnerability_item['id']
                            if db.session.query(CommonVulnerabilitiesAndExposures.id).filter_by(CVEId = reprow_cveId).first() is None:
                                my_cve = CommonVulnerabilitiesAndExposures(CVEId=reprow_cveId)
                                db.session.add(my_cve)
                            else:
                                my_cve = db.session.query(CommonVulnerabilitiesAndExposures).filter_by(CVEId=reprow_cveId).one()

                            if VulnerabilityReport.query.join(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter((VulnerabilityReportVulnerabilitiesLink.vreport_id == my_json_report.id) & (VulnerabilityReportVulnerabilitiesLink.cve_id == my_cve.id)).first() is None:
                                my_link = VulnerabilityReportVulnerabilitiesLink(vreport_id=my_json_report.id, cve_id=my_cve.id)
                                db.session.add(my_link)
                            else:
                                my_link = VulnerabilityReport.query.join(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter((VulnerabilityReportVulnerabilitiesLink.vreport_id == my_json_report.id) & (VulnerabilityReportVulnerabilitiesLink.cve_id == my_cve.id)).first()
                            my_link.asset_id = my_repo_asset.id
                            my_link.VReport_assetID = my_asset_IP if my_asset_IP is not None else ""
                            my_link.VReport_assetIp = my_asset_IP if my_asset_IP is not None else ""
                            my_link.VReport_port = item['port'] if item['port'] is not None else ""
                            my_link.VReport_CVSS_score = vulnerability_item['cvss'] if vulnerability_item['cvss'] is not None else ""
                            my_link.comments = vulnerability_item['is_exploit'] if vulnerability_item['is_exploit'] is not None else ""
                            try:
                                db.session.commit()
                                # flash('Vulnerability "" Added Succesfully')
                                # flash('Vulnerability "{}" Added Succesfully'.format(my_link.cve_id))
                            except SQLAlchemyError as e:
                                db.session.rollback()
                                continue
                            # update_cve_scores(reprow_cveId)
                            # TODO: It's not needed to call the update CVE and CWE functions

            else:
                continue
        return 1

# region Call NVD API to update CVE scores and get CWEs
# region Call NVD API to update CVE scores
def update_cve_scores(cveId):
    response = requests.get("https://services.nvd.nist.gov/rest/json/cve/1.0/" + cveId)
    if response is not None and response.status_code == 200:
        NVDreport = response.json()
        # update CVE table with API values
        impact = NVDreport['result']['CVE_Items'][0]['impact']
        my_cve = db.session.query(CommonVulnerabilitiesAndExposures).filter_by(CVEId=cveId).one()
        if impact['baseMetricV3']  is not None:
            my_cve.exploitabilityScore = impact['baseMetricV3']['exploitabilityScore'] if impact['baseMetricV3']['exploitabilityScore'] is not None else ""
            my_cve.impactScore = impact['baseMetricV3']['impactScore'] if impact['baseMetricV3']['impactScore'] is not None else ""
            my_cve.accessVector = impact['baseMetricV3']['cvssV3']['attackVector'] if impact[
                                                                                          'baseMetricV3'][
                                                                                          'cvssV3'][
                                                                                          'attackVector'] is not None else ""
            my_cve.accessComplexity = impact['baseMetricV3']['cvssV3']['attackComplexity'] if impact[
                                                                                                  'baseMetricV3'][
                                                                                                  'cvssV3'][
                                                                                                  'attackComplexity'] is not None else ""
            my_cve.obtainAllPrivilege = impact['baseMetricV3']['cvssV3']['obtainAllPrivilege'] if impact['baseMetricV3']['cvssV3']['obtainAllPrivilege'] is not None else ""
            my_cve.userInteractionRequired = impact['baseMetricV3']['cvssV3']['userInteractionRequired'] if impact['baseMetricV3']['cvssV3']['userInteractionRequired'] is not None else ""
            my_cve.confidentialityImpact = impact['baseMetricV3']['cvssV3']['confidentialityImpact'] if impact['baseMetricV3']['cvssV3']['confidentialityImpact'] is not None else ""
            my_cve.integrityImpact = impact['baseMetricV3']['cvssV3']['integrityImpact'] if impact['baseMetricV3']['cvssV3']['integrityImpact'] is not None else ""
            my_cve.availabilityImpact = impact['baseMetricV3']['cvssV3']['availabilityImpact'] if impact['baseMetricV3']['cvssV3']['availabilityImpact'] is not None else ""
            my_cve.baseScore = impact['baseMetricV3']['cvssV3']['baseScore'] if impact['baseMetricV3']['cvssV3']['baseScore'] is not None else ""
            my_cve.severity = impact['baseMetricV3']['cvssV3']['baseSeverity'] if impact['baseMetricV3']['cvssV3'][
                                                                                      'baseSeverity'] is not None else ""
        else:
            my_cve.severity = impact['baseMetricV2']['severity'] if impact['baseMetricV2']['severity'] is not None else ""
            my_cve.exploitabilityScore = impact['baseMetricV2']['exploitabilityScore'] if \
                impact['baseMetricV2'][
                    'exploitabilityScore'] is not None else ""
            my_cve.impactScore = impact['baseMetricV2']['impactScore'] if impact["baseMetricV2"][
                                                                              'impactScore'] is not None else ""
            my_cve.obtainAllPrivilege = impact['baseMetricV2']['obtainAllPrivilege'] if \
                impact['baseMetricV2'][
                    'obtainAllPrivilege'] is not None else ""
            my_cve.obtainUserPrivilege = impact["baseMetricV2"]['obtainUserPrivilege'] if \
                impact['baseMetricV2'][
                    'obtainUserPrivilege'] is not None else ""
            my_cve.obtainOtherPrivilege = impact["baseMetricV2"]['obtainOtherPrivilege'] if \
                impact['baseMetricV2'][
                    'obtainOtherPrivilege'] is not None else ""
            my_cve.userInteractionRequired = impact["baseMetricV2"]['userInteractionRequired'] if \
                impact['baseMetricV2'][
                    'userInteractionRequired'] is not None else ""
            my_cve.accessVector = impact['baseMetricV2']['cvssV2']['accessVector'] if impact[
                                                                                          'baseMetricV2'][
                                                                                          'cvssV2'][
                                                                                          'accessVector'] is not None else ""
            my_cve.accessComplexity = impact['baseMetricV2']['cvssV2']['accessComplexity'] if impact[
                                                                                                  'baseMetricV2'][
                                                                                                  'cvssV2'][
                                                                                                  'accessComplexity'] is not None else ""
            my_cve.authentication = impact['baseMetricV2']['cvssV2']['authentication'] if impact[
                                                                                              'baseMetricV2'][
                                                                                              'cvssV2'][
                                                                                              'authentication'] is not None else ""
            my_cve.confidentialityImpact = impact['baseMetricV2']['cvssV2']['confidentialityImpact'] if \
                impact[
                    'baseMetricV2']['cvssV2']['confidentialityImpact'] is not None else ""
            my_cve.integrityImpact = impact['baseMetricV2']['cvssV2']['integrityImpact'] if impact[
                                                                                                'baseMetricV2'][
                                                                                                'cvssV2'][
                                                                                                'integrityImpact'] is not None else ""
            my_cve.availabilityImpact = impact['baseMetricV2']['cvssV2']['availabilityImpact'] if impact[
                                                                                                      'baseMetricV2'][
                                                                                                      'cvssV2'][
                                                                                                      'availabilityImpact'] is not None else ""
            my_cve.baseScore = impact['baseMetricV2']['cvssV2']['baseScore'] if impact['baseMetricV2']['cvssV2']['baseScore'] is not None else ""
        db.session.add(my_cve)

        # Get CWEs and link them with CVE
        for api_cve_desc, api_cveId, api_cweId in get_cwe_codes_from_API_report(NVDreport):
            my_cve.description = api_cve_desc
            cwe_number = api_cweId.split("CWE-", 1)[1].strip()
            if cwe_number.isnumeric():
                if db.session.query(
                        exists().where(CommonWeaknessEnumeration.CWEId == cwe_number)).scalar():
                    my_CWE_row = \
                        db.session.query(CommonWeaknessEnumeration).filter_by(CWEId=cwe_number)[0]
                else:
                    my_CWE_row = CommonWeaknessEnumeration(CWEId=cwe_number)
                    db.session.add(my_CWE_row)
                if db.session.query(CommonVulnerabilitiesAndExposures).filter(
                        VulnerabilitiesWeaknessLink.cwe_id == my_CWE_row.id,
                        VulnerabilitiesWeaknessLink.cve_id == my_cve.id).first() is None:
                    my_cVecWe = VulnerabilitiesWeaknessLink(cve_id=my_cve.id, cwe_id=my_CWE_row.id,
                                                            date=datetime.utcnow())
                else:
                    my_cVecWe = db.session.query(CommonVulnerabilitiesAndExposures).filter(
                        VulnerabilitiesWeaknessLink.cwe_id == my_CWE_row.id,
                        VulnerabilitiesWeaknessLink.cve_id == my_cve.id).first()
                    my_cVecWe.date = datetime.utcnow()
                db.session.add(my_cVecWe)
        try:
            db.session.commit()
        except SQLAlchemyError as e:
            db.session.rollback()
            return -1
        return 1
# endregion

# region NVD API get_cwe_codes
def get_cwe_codes_from_API_report(APIreport):
    for item in APIreport['result']['CVE_Items']:
        itemdescr = item["cve"]["description"]["description_data"][0]["value"]
        for problem in item["cve"]['problemtype']['problemtype_data']:
            for descr in problem['description']:
                yield itemdescr, item["cve"]["CVE_data_meta"]["ID"], descr["value"]

# endregion CWE
# endregion CVE & CWE
# endregion VAaaS report

# region Insert information from Certification Report
def certification_report_json(report_details):
    """ report_details should be a json object"""
    # obj = json.loads(report_details)
    if report_details["id"] is not None:
        reprow_reportId = report_details["id"]
        if db.session.query(exists().where(VulnerabilityReport.reportId == reprow_reportId)).scalar():
            my_json_report = db.session.query(VulnerabilityReport).filter_by(reportId=reprow_reportId).one()
        else:
            my_json_report = VulnerabilityReport(reportId=reprow_reportId)
        my_json_report.scan_start_time = report_details["start"] if report_details["start"] is not None else ""
        my_json_report.scan_end_time = report_details["end"] if report_details["end"] is not None else ""
        my_json_report.source_component = 2
        my_json_report.source_attackType = report_details["attackType"] if report_details["attackType"] is not None else ""
        my_json_report.source_eventsCount = report_details["eventsCount"] if report_details["eventsCount"] is not None else ""
        my_json_report.source_riskScore = report_details["riskScore"] if report_details["riskScore"] is not None else ""
        my_json_report.source_severity = report_details["severity"] if report_details["severity"] is not None else ""
        db.session.add(my_json_report)
        try:
            db.session.commit()
            # flash('Vulnerability from ACS Report "{}" Added Succesfully'.format(my_json_report.reportId))
        except SQLAlchemyError as e:
            db.session.rollback()
            return -1

        my_asset_IP = None
        my_asset_Name = None
        for item in report_details['data']:
            # Get asset IP
            if item['agent.ip'] is not None:
                # We do not know if this report includes only one agent
                my_asset_IP = item['agent.ip']
                my_asset_Name = item['agent.name']
                if db.session.query(RepoAsset.id).filter_by(ip=my_asset_IP).first() is None:
                    my_repo_asset = RepoAsset(ip=my_asset_IP)
                    my_repo_asset.name = my_asset_Name if my_asset_Name is not None else ""
                    db.session.add(my_repo_asset)
                    try:
                        db.session.commit()
                        # flash('Asset "{}" Added Succesfully'.format(my_repo_asset.ip))

                        # TODO: Send alert for the new Asset to the EndUser
                    except SQLAlchemyError as e:
                        db.session.rollback()
                    send_alert_new_asset(my_repo_asset.id)
                else:
                    my_repo_asset = db.session.query(RepoAsset).filter_by(ip=my_asset_IP).first()

                # Get CVE from the result nodes of the report
                reprow_cve = item['rule.description']
                reprow_cveId = reprow_cve.split(' ')[0]
                if db.session.query(CommonVulnerabilitiesAndExposures.id).filter_by(CVEId=reprow_cveId).first() is None:
                    my_cve = CommonVulnerabilitiesAndExposures(CVEId=reprow_cveId)
                    db.session.add(my_cve)
                else:
                    my_cve = db.session.query(CommonVulnerabilitiesAndExposures).filter_by(CVEId=reprow_cveId).first()

                if VulnerabilityReport.query.join(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter((VulnerabilityReportVulnerabilitiesLink.vreport_id == my_json_report.id) & (VulnerabilityReportVulnerabilitiesLink.cve_id == my_cve.id)).first() is None:
                    my_link = VulnerabilityReportVulnerabilitiesLink(vreport_id=my_json_report.id, cve_id=my_cve.id)
                    db.session.add(my_link)
                else:
                    my_link = VulnerabilityReport.query.join(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter((VulnerabilityReportVulnerabilitiesLink.vreport_id == my_json_report.id) & (VulnerabilityReportVulnerabilitiesLink.cve_id == my_cve.id)).first()
                my_link.asset_id = my_repo_asset.id
                my_link.VReport_assetIp = item['agent.ip'] if item['agent.ip'] is not None else ""
                my_link.VReport_CVSS_score = '{0:.2f}'.format(float(item['cvss3'])) if item['cvss3'] is not None else ""
                my_link.comments = item['rule.description'] if item['rule.description'] is not None else ""

                # Connect the result nodes of the report with the relevant Threat
                reprow_threat = report_details['attackType']
                if db.session.query(RepoThreat.id).filter_by(name=reprow_threat).first() is not None:
                    my_threat = db.session.query(RepoThreat).filter_by(name=reprow_threat).first()
                    if db.session.query(repo_threat_common_vulnerabilities_and_exposures_association_table).filter_by(repo_threat_id=my_threat.id, common_vulnerabilities_and_exposures_id=my_cve.id).first() is None:
                        my_threat.cves.append(my_cve)

                try:
                    db.session.commit()
                except SQLAlchemyError as e:
                    db.session.rollback()
                    continue
            else:
                continue
        return 1
# endregion Insert information from Certification Report

# region Insert Asset information from DTM
def getAssetsfromDTM(report_details):
    """ report_details should be a json object"""
    # obj = json.loads(report_details)
    if report_details["ip"] is not None:
        reprow_dtm_asset_id = report_details["ip"]
        if db.session.query(RepoAsset.id).filter_by(ip = reprow_dtm_asset_id).first() is not None:
            my_db_asset = db.session.query(RepoAsset).filter_by(ip=reprow_dtm_asset_id).first()
        else:
            my_db_asset = RepoAsset(ip=reprow_dtm_asset_id)
        my_db_asset.mac_address = report_details["physicalAddress"] if report_details["physicalAddress"] is not None else ""
        my_db_asset.last_touch_date = report_details["lastTouch"] if report_details["lastTouch"] is not None else ""
        db.session.add(my_db_asset)
        try:
            db.session.commit()
            # flash('Asset "{}" Added Succesfully'.format(my_db_asset.ip))
        except SQLAlchemyError as e:
            db.session.rollback()
            return -1
        send_alert_new_asset(my_db_asset.id)
    #     TODO FLAG TO ENSURE CREATE IS CALLED ONLY ON CREATE NOT UPDATE WHICH IS FIRST IF
    return 1

# endregion Insert Asset information from DTM

# region Handle SIEM alerts
def siem_alerts(report_details):
    if report_details["attackType"] is not None:
        # if report_details["threat"] is not None: Retrieve Threat id based on threat_name
        alert_threat = report_details['attackType']
        if db.session.query(RepoThreat.id).filter_by(name=alert_threat).first() is not None:
            my_threat = db.session.query(RepoThreat).filter_by(name=alert_threat).first()
            alert_asset_ip = report_details['agent.ip']
            if db.session.query(RepoAsset.ip).filter_by(ip=alert_asset_ip).first() is not None:
                my_asset = db.session.query(RepoAsset).filter_by(ip=alert_asset_ip).first()

                print("_______MYASSETIS____________")
                print(my_asset)
                print("_______MYTHREATIS____________")
                print(my_threat)
                this_risk_assessment = RepoRiskAssessment.query.filter_by(repo_threat_id=my_threat.id,
                                                                          repo_asset_id=my_asset.id).first()
                if this_risk_assessment is None:
                    pass
                else:
                    print("HELLO 1")
                    risk_assessment_result = start_risk_assessment_alert(my_threat.id, my_asset.id,
                                                                         materialisation_value=100,
                                                                         consequence_values=100)
                    print("HELLO 1.5")
                    risk_assessment_saved = risk_assessment_save_report(my_threat.id, my_asset.id,
                                                                        risk_assessment_result, "incident")
                    send_risk_report(risk_assessment_saved.id, my_asset.id, my_threat.id)
                    print("HELLO 2")
                    security_event_risk_reports(risk_assessment_saved.id)

                print("HELLO 3")
                # TODO: Initiate Risk Assessment for this asset with mat1 and the rest nodes = 100% [we will define these for each threat- in RA call]
                other_net_assets = get_all_assets_of_network_group(my_asset)
                print("---- OTHER ASSETS ARE ----")
                print(other_net_assets)
                if other_net_assets:
                    print("Assets on this Networks")
                    for each_asset in other_net_assets:
                        # Check if the asset and threat pair has a risk assessment report ready
                        # If there isnt continue
                        this_risk_assessment = RepoRiskAssessment.query.filter_by(repo_threat_id=my_threat.id,
                                                                                  repo_asset_id=each_asset.id).first()


                        if not each_asset.verified:
                            print("I'm not verified")
                            send_alert_new_asset(each_asset.id)
                            continue
                            # risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id, materialisation_value=100,consequence_values=100)
                            # risk_assessment_saved = risk_assessment_save_report(my_threat.id, my_asset.id,
                            #                                                     risk_assessment_result, "incident_secondary")
                        else:
                            if this_risk_assessment is None:
                                # TODO SEND ALERT TO UPDATE INFORMATION FOR THIS ASSET THREAT PAIR TO CONDUCT RISK ASSESSMENT
                                send_alert_info_update_needed(asset_id=each_asset.id, threat_id=my_threat.id,
                                                              threat_exposure_info=0,
                                                              threat_materialisation_info=0, threat_impact_info=0,
                                                              objective_info=0,
                                                              utility_info=0)
                                continue

                            asset_reputation_value = get_asset_reputation(each_asset.common_id)
                            print(asset_reputation_value)
                            asset_vulnerability_value = get_asset_vulnerabilities_status(each_asset.id, my_threat.id)
                            print(asset_vulnerability_value)
                            if each_asset.type_fk == my_asset.type_fk:
                                print("Asset of the same type on the same network: {0}, type: {1}".format(each_asset.name, each_asset.type_fk))
                                if (asset_vulnerability_value[0] >= 7.5) or ((asset_vulnerability_value[0] + asset_vulnerability_value[1])/2 >= 7.5):
                                    print("Over 7.5 : {0}".format(asset_vulnerability_value[0]))
                                    risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id, materialisation_value=100)
                                    risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                        risk_assessment_result, "incident_secondary")
                                    send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                    security_event_risk_reports(risk_assessment_saved.id)
                                    # TODO: Initiate Risk Assessment for these assets with mat1 = 100%
                                elif 5 <= asset_vulnerability_value[0] < 7.5:
                                    print("5 to 7.5 the average: {0}".format(asset_vulnerability_value[0]))
                                    risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id, materialisation_value_increase=asset_vulnerability_value[0]*10)
                                    risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                        risk_assessment_result, "incident_secondary")
                                    send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                    security_event_risk_reports(risk_assessment_saved.id)
                                    # TODO: Initiate Risk Assessment for these assets with mat1 = mat1 * (1+ asset_vulnerability_value[0]/10)
                                    #   obviously it should be <=100%
                                else:
                                    if asset_reputation_value < 100 and each_asset.value == 3 and asset_reputation_value != -1:
                                        print("Reputation <100 and Asset value =3: {0} - {1}".format(asset_reputation_value, each_asset.value))
                                        risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                    materialisation_value_increase=100)
                                        risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                            risk_assessment_result, "incident_secondary")
                                        send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                        security_event_risk_reports(risk_assessment_saved.id)
                                        # TODO: Initiate Risk Assessment for these assets with exposure = 100%
                                    elif asset_reputation_value < 100 and each_asset.value == 2 and asset_reputation_value != -1:
                                        print("Reputation <100 and Asset value =2: {0} - {1}".format(asset_reputation_value, each_asset.value))
                                        risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id, exposure_value_increase=20)
                                        risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                            risk_assessment_result, "incident_secondary")
                                        send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                        security_event_risk_reports(risk_assessment_saved.id)
                                        # TODO: Initiate Risk Assessment for these assets with [exposure = exposure * 1,2]  obviously <=100%
                                    else:
                                        print("Other: {0} - {1}".format(asset_reputation_value, each_asset.value))
                                        risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id, exposure_value_increase=10)
                                        risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                            risk_assessment_result, "incident_secondary")
                                        send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                        security_event_risk_reports(risk_assessment_saved.id)
                                        # TODO: Initiate Risk Assessment for these assets with [exposure = exposure * 1,1]  obviously <=100%

                            else:
                                print("Asset of different type on the same network: {0}, type: {1}".format(each_asset.name, each_asset.type_fk))
                                if (asset_vulnerability_value[0] >= 7.5) or ((asset_vulnerability_value[0] + asset_vulnerability_value[1])/2 >= 7.5):
                                    print("Other type of Asset - Over 7.5 : {0}".format(asset_vulnerability_value[0]))
                                    risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                materialisation_value_increase=asset_vulnerability_value[0]*10)
                                    risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                        risk_assessment_result,
                                                                                        "incident_secondary")
                                    send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                    security_event_risk_reports(risk_assessment_saved.id)
                                    # TODO: Initiate Risk Assessment for these assets with mat1 = mat1 * (1+ asset_vulnerability_value[0]/10)
                                    #   obviously it should be <=100%
                                else:
                                    if asset_reputation_value < 100 and each_asset.value == 3 and asset_reputation_value != -1:
                                        print("Reputation <100 and Asset value =3: {0} - {1}".format(
                                            asset_reputation_value, each_asset.value))
                                        risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                    exposure_value=100)
                                        risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                           risk_assessment_result,
                                                                                           "incident_secondary")
                                        send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                        security_event_risk_reports(risk_assessment_saved.id)
                                        # TODO: Initiate Risk Assessment for these assets with exposure = 100%
                                    elif asset_reputation_value < 100 and each_asset.value == 2 and asset_reputation_value != -1:
                                        print("Reputation <100 and Asset value =2: {0} - {1}".format(
                                            asset_reputation_value, each_asset.value))
                                        risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                    exposure_value_increase=20)
                                        risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                            risk_assessment_result,
                                                                                            "incident_secondary")
                                        send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                        security_event_risk_reports(risk_assessment_saved.id)
                                        # TODO: Initiate Risk Assessment for these assets with [exposure = exposure * 1,2]  obviously <=100%
                                    else:
                                        risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                    exposure_value_increase=10)
                                        risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                            risk_assessment_result,
                                                                                            "incident_secondary")
                                        print("Other: {0} - {1}".format(asset_reputation_value, each_asset.value))
                                        send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                        security_event_risk_reports(risk_assessment_saved.id)
                                        # TODO: Initiate Risk Assessment for these assets with [exposure = exposure * 1,1]  obviously <=100%
                else:
                    print("Assets on other Networks")
                    # Check if the asset and threat pair has a risk assessment report ready
                    # If there isnt continue
                    assets_not_on_net = get_assets_not_on_netgroup_with_threat_vuln(my_asset, my_threat.id)
                    for each_asset in assets_not_on_net:
                        this_risk_assessment = RepoRiskAssessment.query.filter_by(repo_threat_id=my_threat.id,
                                                                                  repo_asset_id=each_asset.id).first()


                        if not each_asset.verified:
                            print("I'm not verified")
                            send_alert_new_asset(each_asset.id)
                            continue
                        else:
                            if this_risk_assessment is None:
                                send_alert_info_update_needed(asset_id=each_asset.id, threat_id=my_threat.id,
                                                              threat_exposure_info=0,
                                                              threat_materialisation_info=0, threat_impact_info=0,
                                                              objective_info=0,
                                                              utility_info=0)
                                continue

                            asset_reputation_value = get_asset_reputation(each_asset.common_id)
                            print(asset_reputation_value)
                            asset_vulnerability_value = get_asset_vulnerabilities_status(each_asset.id, my_threat.id)
                            print(asset_vulnerability_value)
                            print("Asset on different network: {0}, type: {1}".format(each_asset.name, each_asset.type_fk))
                            if (asset_vulnerability_value[0] >= 5) or ((asset_vulnerability_value[0] + asset_vulnerability_value[1])/2 >= 7.5):
                                print("Over 5 and average 7.5 : {0}".format(asset_vulnerability_value[0]))
                                risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                                     materialisation_value_increase=asset_vulnerability_value[0]*10)
                                risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                    risk_assessment_result,
                                                                                    "incident_secondary")
                                send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                security_event_risk_reports(risk_assessment_saved.id)
                                # TODO: Initiate Risk Assessment for these assets with mat1 = mat1 * (1+ asset_vulnerability_value[0]/10)
                                #   obviously it should be <=100%
                            else:
                                if asset_reputation_value < 100 and each_asset.value == 3 and asset_reputation_value != -1:
                                    print("Reputation <100 and Asset value =3: {0} - {1}".format(asset_reputation_value, each_asset.value))
                                    risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                                         exposure_value=100
                                                                                         )
                                    risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                        risk_assessment_result,
                                                                                        "incident_secondary")
                                    send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                    security_event_risk_reports(risk_assessment_saved.id)
                                    # TODO: Initiate Risk Assessment for these assets with exposure = 100%
                                elif asset_reputation_value < 100 and each_asset.value == 2 and asset_reputation_value != -1:
                                    print("Reputation <100 and Asset value =2: {0} - {1}".format(asset_reputation_value, each_asset.value))
                                    risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                                         exposure_value_increase=20
                                                                                         )
                                    risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                        risk_assessment_result,
                                                                                        "incident_secondary")
                                    send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                    security_event_risk_reports(risk_assessment_saved.id)
                                    # TODO: Initiate Risk Assessment for these assets with [exposure = exposure * 1,2]  obviously <=100%
                                else:
                                    print("Other: {0} - {1}".format(asset_reputation_value, each_asset.value))
                                    risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                                         exposure_value_increase=10
                                                                                         )
                                    risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                        risk_assessment_result,
                                                                                        "incident_secondary")
                                    send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                    security_event_risk_reports(risk_assessment_saved.id)
                                    # TODO: Initiate Risk Assessment for these assets with [exposure = exposure * 1,1]  obviously <=100%
            else:
                print("No such asset")
                assets_not_on_net = get_assets_not_on_netgroup_with_threat_vuln('', my_threat.id)
                for each_asset in assets_not_on_net:
                    if not each_asset.verified:
                        print("I'm not verified")
                        # TODO: Initiate Risk Assessment for this asset with mat1 and the rest nodes = 100% [we will define these for each threat- in RA call]
                        #      if we have enough data, otherwise WHAT???
                    else:
                        asset_reputation_value = get_asset_reputation(each_asset.common_id)
                        print(asset_reputation_value)
                        asset_vulnerability_value = get_asset_vulnerabilities_status(each_asset.id, my_threat.id)
                        print(asset_vulnerability_value)
                        print("Asset on different network: {0}, type: {1}".format(each_asset.name, each_asset.type_fk))
                        if (asset_vulnerability_value[0] >= 5) or (
                                (asset_vulnerability_value[0] + asset_vulnerability_value[1]) / 2 >= 7.5):
                            print("Over 5 and average 7.5 : {0}".format(asset_vulnerability_value[0]))
                            risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                                 materialisation_value_increase=asset_vulnerability_value[0]*10
                                                                                 )
                            risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                risk_assessment_result,
                                                                                "incident_secondary")
                            send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                            security_event_risk_reports(risk_assessment_saved.id)
                            # TODO: Initiate Risk Assessment for these assets with mat1 = mat1 * (1+ asset_vulnerability_value[0]/10)
                            #   obviously it should be <=100%
                        else:
                            if asset_reputation_value < 100 and each_asset.value == 3 and asset_reputation_value != -1:
                                print("Reputation <100 and Asset value =3: {0} - {1}".format(asset_reputation_value,
                                                                                             each_asset.value))
                                risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                                     exposure_value= 100
                                                                                     )
                                risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                    risk_assessment_result,
                                                                                    "incident_secondary")
                                send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                security_event_risk_reports(risk_assessment_saved.id)
                                # TODO: Initiate Risk Assessment for these assets with exposure = 100%
                            elif asset_reputation_value < 100 and each_asset.value == 2 and asset_reputation_value != -1:
                                print("Reputation <100 and Asset value =2: {0} - {1}".format(asset_reputation_value,
                                                                                             each_asset.value))
                                risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                                     exposure_value_increase=20
                                                                                     )
                                risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                    risk_assessment_result,
                                                                                    "incident_secondary")
                                send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                security_event_risk_reports(risk_assessment_saved.id)
                                # TODO: Initiate Risk Assessment for these assets with [exposure = exposure * 1,2]  obviously <=100%
                            else:
                                print("Other: {0} - {1}".format(asset_reputation_value, each_asset.value))
                                risk_assessment_result = start_risk_assessment_alert(my_threat.id, each_asset.id,
                                                                                     exposure_value=10
                                                                                     )
                                risk_assessment_saved = risk_assessment_save_report(my_threat.id, each_asset.id,
                                                                                    risk_assessment_result,
                                                                                    "incident_secondary")
                                send_risk_report(risk_assessment_saved.id, each_asset.id, my_threat.id)
                                security_event_risk_reports(risk_assessment_saved.id)
                                # TODO: Initiate Risk Assessment for these assets with [exposure = exposure * 1,1]  obviously <=100%
        else:   # If we can not identify the threat then do nothing
            return -1


def get_all_assets_of_network_group(base_asset):
    if base_asset.net_group_fk is not None:
        if (db.session.query(RepoAsset).filter(RepoAsset.id != base_asset.id,
                                                                       RepoAsset.net_group_fk == base_asset.net_group_fk)) is not None:
            other_assets = db.session.query(RepoAsset).filter(RepoAsset.id != base_asset.id,
                                                              RepoAsset.net_group_fk == base_asset.net_group_fk).all()
            return other_assets
        else:
            return []
    else:
        return []


def get_assets_not_on_netgroup_with_threat_vuln(base_asset, threat_id):
    if threat_id is None:
        return []
    # print("Get list of relevant vulnerabilities to this threat----------")
    if db.session.query(repo_threat_common_vulnerabilities_and_exposures_association_table).filter_by(
            repo_threat_id=threat_id).first() is not None:
        vuln_threat_assoc = db.session.query(
            repo_threat_common_vulnerabilities_and_exposures_association_table).filter_by(
            repo_threat_id=threat_id).all()
        vuln_threat = []
        for item in vuln_threat_assoc:
            vuln_threat.append(item[1])
    else:
        vuln_threat = []

    # print("Get list of relevant assets based on the above vulnerabilities ----------")
    if db.session.query(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter(CommonVulnerabilitiesAndExposures.id.in_(vuln_threat)).first() is not None:
        vulnerabilities_threat_all = db.session.query(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter(CommonVulnerabilitiesAndExposures.id.in_(vuln_threat)).all()
        other_assets = []
        for item in vulnerabilities_threat_all:
            other_assets.append(item.asset_id)
    else:
        other_assets = []

    # print("Get list of relevant assets based on the above vulnerabilities ----------")
    if (db.session.query(RepoAsset).filter(RepoAsset.id.in_(other_assets),
                                               RepoAsset.net_group_fk != base_asset.net_group_fk)) is not None:
        other_assets_assoc = db.session.query(RepoAsset).filter(RepoAsset.id.in_(other_assets), RepoAsset.net_group_fk != base_asset.net_group_fk).all()
        output_assets = []
        for item in other_assets_assoc:
            output_assets.append(item.asset_id)
        # print(other_assets)
    else:
        output_assets = []

    return output_assets


def get_asset_reputation(asset_common_id):
    if asset_common_id is None:
        return -1
    if db.session.query(RepoAssetReputation.reputation).filter_by(global_asset_id=asset_common_id).first() is not None:
        reputation_value = db.session.query(RepoAssetReputation).filter_by(global_asset_id=asset_common_id).order_by(RepoAssetReputation.id.desc()).first()
        # We take the last entry for this asset
        return reputation_value.reputation if reputation_value.reputation is not None else -1
    else:
        return -1

def get_asset_vulnerabilities_status(asset_id, threat_id):
    if asset_id is None:
        return []
    reports = []
    # print("Get most recent reports---------------")
    if db.session.query(VulnerabilityReport.id).join(VulnerabilityReportVulnerabilitiesLink).filter(VulnerabilityReport.source_component == 1).filter(VulnerabilityReportVulnerabilitiesLink.asset_id == asset_id).order_by(VulnerabilityReport.id.desc()).first() is not None:
        vaaas_report = db.session.query(VulnerabilityReport).join(VulnerabilityReportVulnerabilitiesLink).filter(VulnerabilityReport.source_component == 1).filter(VulnerabilityReportVulnerabilitiesLink.asset_id == asset_id).order_by(VulnerabilityReport.id.desc()).first()
        reports.append(vaaas_report.id)

    if db.session.query(VulnerabilityReport.id).join(VulnerabilityReportVulnerabilitiesLink).filter(VulnerabilityReport.source_component == 2).filter(VulnerabilityReportVulnerabilitiesLink.asset_id == asset_id).order_by(VulnerabilityReport.id.desc()).first() is not None:
        acs_report = db.session.query(VulnerabilityReport).join(VulnerabilityReportVulnerabilitiesLink).filter(VulnerabilityReport.source_component == 2).filter(VulnerabilityReportVulnerabilitiesLink.asset_id == asset_id).order_by(VulnerabilityReport.id.desc()).first()
        reports.append(acs_report.id)
    # for item in reports:
    #     print(item)
    # print("-----")
    # print("Get list of relevant vulnerabilities----------")
    if db.session.query(repo_threat_common_vulnerabilities_and_exposures_association_table).filter_by(repo_threat_id=threat_id).first() is not None:
        vuln_threat_assoc = db.session.query(repo_threat_common_vulnerabilities_and_exposures_association_table).filter_by(repo_threat_id=threat_id).all()
        vuln_threat=[]
        for item in vuln_threat_assoc:
            vuln_threat.append(item[1])
        # print(vuln_threat)
    else:
        vuln_threat = []
    vulnerabilities_threat_all = db.session.query(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter(VulnerabilityReportVulnerabilitiesLink.vreport_id.in_(reports)).filter(CommonVulnerabilitiesAndExposures.id.in_(vuln_threat)).all()
    # print(" Threat is associated with these vuln:------------")
    relevant_vuln_score = 0
    for item in vulnerabilities_threat_all:
        if db.session.query(RepoControl.id).filter_by(vulnerability_id=item.id).first() is not None:
            max_control_effectiveness = db.session.query(func.max(RepoControl.effectiveness)).filter_by(vulnerability_id=item.id).scalar()
        else:
            max_control_effectiveness = 0
        vuln_score = float(item.VReport_CVSS_score)*(1 - max_control_effectiveness/100) if float(item.VReport_CVSS_score)*(1 - max_control_effectiveness/100) is not None else 0
        relevant_vuln_score = max(relevant_vuln_score, vuln_score)
        # print(item.id, item.comments)
        # print(vuln_score)
    # print("Max Relevant Vulnerability score: {0}".format(relevant_vuln_score))
    # print("------------")
    vulnerabilities_rest_all = db.session.query(VulnerabilityReportVulnerabilitiesLink).join(CommonVulnerabilitiesAndExposures).filter(VulnerabilityReportVulnerabilitiesLink.vreport_id.in_(reports)).filter(CommonVulnerabilitiesAndExposures.id.notin_(vuln_threat)).all()
    # print("Threat is Not associated with these vuln:------------")
    rest_vuln_score = 0
    for item in vulnerabilities_rest_all:
        if db.session.query(RepoControl.id).filter_by(vulnerability_id=item.id).first() is not None:
            max_control_effectiveness = db.session.query(func.max(RepoControl.effectiveness)).filter_by(vulnerability_id=item.id).scalar()
        else:
            max_control_effectiveness = 0
        vuln_score = float(item.VReport_CVSS_score) * (1 - max_control_effectiveness / 100) if float(
            item.VReport_CVSS_score) * (1 - max_control_effectiveness / 100) is not None else 0
        rest_vuln_score = max(rest_vuln_score, vuln_score)
        # print(item.id, item.comments)
        # print(vuln_score)
    # print("Max Rest Vulnerability score: {0}".format(rest_vuln_score))

    return [relevant_vuln_score, rest_vuln_score]

    #publish to kafka everything

# endregion Handle SIEM alerts

# region OLD functions
# region get_assets
# temporarily from VaasReport table
def get_assets():
    if db.session.query(VulnerabilityReportVulnerabilitiesLink).distinct(
            VulnerabilityReportVulnerabilitiesLink.VReport_assetID).count() > 0:
        list_of_assets = db.session.query(VulnerabilityReportVulnerabilitiesLink).distinct(
            VulnerabilityReportVulnerabilitiesLink.VReport_assetID)
        return list_of_assets
    else:
        return []
    # db.session.query(your_table.column1.distinct()).filter_by(column2 = 'some_column2_value').all()


#
# def get_assetsfromrepository():
#     if db.session.query().distinct(HardwareAsset.id).count() > 0:
#         list_of_assets = db.session.query(HardwareAsset).distinct(HardwareAsset.id)
#         return list_of_assets
#     else:
#         return -1

# endregion

# region get Recommended CVEs for an asset
def get_cve_recommendations(asset_id):
    if db.session.query(VulnerabilityReportVulnerabilitiesLink).distinct(
            VulnerabilityReportVulnerabilitiesLink.cve_id).filter(
        VulnerabilityReportVulnerabilitiesLink.VReport_assetID == asset_id).count() > 0:
        list_of_cve = db.session.query(VulnerabilityReportVulnerabilitiesLink.cve_id).distinct(
            VulnerabilityReportVulnerabilitiesLink.cve_id).filter(
            VulnerabilityReportVulnerabilitiesLink.VReport_assetID == asset_id)
        result = db.session.query(CommonVulnerabilitiesAndExposures).filter(
            CommonVulnerabilitiesAndExposures.id.in_(list_of_cve))
        return result.all()
    else:
        return -1


# endregion

# region get Recommended CWEs for a selected CVE
def get_cwe_recommendations(selected_cve_id):
    if db.session.query(VulnerabilitiesWeaknessLink).distinct(VulnerabilitiesWeaknessLink.cwe_id).filter(
            VulnerabilitiesWeaknessLink.cve_id == selected_cve_id).count() > 0:
        list_of_cwe = db.session.query(VulnerabilitiesWeaknessLink.cwe_id).distinct(
            VulnerabilitiesWeaknessLink.cwe_id).filter(VulnerabilitiesWeaknessLink.cve_id == selected_cve_id)
        result = db.session.query(CommonWeaknessEnumeration).filter(CommonWeaknessEnumeration.id.in_(list_of_cwe))
        return result.all()
    else:
        return -1


# endregion

# region get Recommended CAPECs for a selected CVE
def get_capec_recommendations(selected_cve_id):
    for cwe_item in get_cwe_recommendations(selected_cve_id):
        capecList = []
        for capec_item in CommonAttackPatternEnumerationClassification.query.filter(
                CommonAttackPatternEnumerationClassification.relatedWeaknesses.like("%" + cwe_item.CWEId + "%")):
            capecList.append(capec_item.id)
    if capecList:
        result = db.session.query(CommonAttackPatternEnumerationClassification).distinct(
            CommonAttackPatternEnumerationClassification.capecId).filter(
            CommonAttackPatternEnumerationClassification.id.in_(capecList)) if db.session.query(
            CommonAttackPatternEnumerationClassification).filter(
            CommonAttackPatternEnumerationClassification.id.in_(capecList)) is not None else -1
        return result.all()
    else:
        return -1


# endregion
# endregion OLD functions

# region test Area
# path_to_VAaaS_report = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)), 'Json_texts', 'report_example_stix.json')
# x = v_report(path_to_VAaaS_report)
#
# print(x)

# path_to_SIEM_report = os.path.join(os.path.abspath(os.path.join(os.path.dirname(__file__), os.path.pardir)),
#                                         'Json_texts', 'siem-certification-vulnerabilities.json')
# with open(path_to_SIEM_report, "r") as fp:
#     obj = json.load(fp)
#     y = certification_report_json(obj)
#     print(y)

# TEST 1
# rep = json.loads('{"attackType":"Data Integrity Violation", "agent.ip":"10.10.50.41"}')
# xx = siem_alerts(rep)
# print(xx)

# TEST 2
# xx = get_asset_reputation(1)
# print(xx)

# TEST 3
# xx = get_asset_vulnerabilities_status(3, 7)
# print(xx)

# send_alert_new_asset(7)
# security_event_risk_reports(8)
# security_event_risk_reports(9)
# security_event_risk_reports(10)
# endregion test Area

