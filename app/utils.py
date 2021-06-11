import subprocess
from copy import deepcopy

from deepdiff import DeepDiff
from flask import jsonify
from sqlalchemy.exc import SQLAlchemyError
from app import db
from app.models import *
# from app.csv_to_json_converter_util import *
from sqlalchemy import exists
from datetime import date, datetime
import openpyxl
import json
import os
import requests
import stix2
import stix2validator
import app.stix2_custom as stix2_custom
# region Insert information from Excel files
# region Insert all CAPEC records from Excel
from app.producer import SendKafkaReport
import pyAgrum as gum


def CAPEC_excel_insertData(capecexcelpath):
    theFile = openpyxl.load_workbook(capecexcelpath)
    currentSheet = theFile.active
    for row in currentSheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # We need to check that the cell is not empty.
            if not db.session.query(
                    exists().where(CommonAttackPatternEnumerationClassification.capecId == row[0])).scalar():
                my_row = CommonAttackPatternEnumerationClassification(capecId=row[0], name=row[1], abstraction=row[2],
                                                                      status=row[3], description=row[4],
                                                                      alternateTerms=row[5], likelihoodOfAttack=row[6],
                                                                      typicalSeverity=row[7],
                                                                      relatedAttackpatterns=row[8],
                                                                      executionFlow=row[9], prerequisites=row[10],
                                                                      skillsRequired=row[11], resourcesRequired=row[12],
                                                                      indicators=row[13], consequences=row[14],
                                                                      mitigations=row[15], exampleInstances=row[16],
                                                                      relatedWeaknesses=row[17],
                                                                      taxonomyMappings=row[18], notes=row[19])
                db.session.add(my_row)
    db.session.commit()
    return 1


# endregion

# region Insert all CWE records from Excel
def CWE_excel_insertData(cweexcelpath):
    theFile = openpyxl.load_workbook(cweexcelpath)
    currentSheet = theFile.active
    for row in currentSheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            row_cweID = str(row[0])
            if db.session.query(exists().where(CommonWeaknessEnumeration.CWEId == row_cweID)).scalar():
                my_row = db.session.query(CommonWeaknessEnumeration).filter_by(CWEId=row_cweID)[0]
            else:
                my_row = CommonWeaknessEnumeration(CWEId=row[0])
            my_row.name = row[1]
            my_row.weakness = row[2]
            my_row.abstraction = row[3]
            my_row.status = row[4]
            my_row.description = row[5]
            my_row.extendedDescription = row[6]
            my_row.relatedWeaknesses = row[7]
            my_row.weaknessOrdinalities = row[8]
            my_row.applicablePlatforms = row[9]
            my_row.backgroundDetails = row[10]
            my_row.alternateTerms = row[11]
            my_row.modesOfIntroduction = row[12]
            my_row.exploitationFactors = row[13]
            my_row.likelihoodOfExploit = row[14]
            my_row.commonConsequences = row[15]
            my_row.detectionMethods = row[16]
            my_row.potentialMitigations = row[17]
            my_row.observedExamples = row[18]
            my_row.functionalAreas = row[19]
            my_row.affectedResources = row[20]
            my_row.taxonomyMappings = row[21]
            my_row.relatedAttackPatterns = row[22]
            my_row.notes = row[23]
            db.session.add(my_row)
    db.session.commit()
    return 1


# endregion

# region Insert all CVE records from Excel
def CVE_excel_insertData(cveexcelpath):
    theFile = openpyxl.load_workbook(cveexcelpath)
    currentSheet = theFile.active
    for row in currentSheet.iter_rows(min_row=2, values_only=True):
        # print(row[0])
        if row[0] is not None:
            if not db.session.query(
                    exists().where(CommonVulnerabilitiesAndExposures.CVEId == row[0])).scalar():
                my_row = CommonVulnerabilitiesAndExposures(CVEId=row[0], status=row[1])
                db.session.add(my_row)
    db.session.commit()
    return 1


# endregion

# region Insert information from VAaaS Report
def v_report(fpath):
    with open(fpath, "r") as fp:
        obj = json.load(fp)
        # todo: check what must be changed if a report has more than one devices
        # print("Test")
        if obj["report"]["@id"] is not None:
            reprow_reportId = obj["report"]["@id"]
            if db.session.query(exists().where(VulnerabilityReport.reportId == reprow_reportId)).scalar():
                my_json_report = db.session.query(VulnerabilityReport).filter_by(reportId=reprow_reportId).one()
            else:
                my_json_report = VulnerabilityReport(reportId=reprow_reportId)
            my_json_report.creation_time = obj["report"]["creation_time"] if obj["report"][
                                                                                 "creation_time"] is not None else ""
            my_json_report.name = obj["report"]["name"] if obj["report"]["name"] is not None else ""
            db.session.add(my_json_report)
            try:
                db.session.commit()
            except SQLAlchemyError as e:
                db.session.rollback()
                return -1
            # Get CVE from the result nodes of the report
            for item in obj['report']['report']['results']['result']:
                print("Test1")
                if item["nvt"]["cve"] == "NOCVE":
                    print("Continue")
                    continue
                else:
                    print("CVE Exists")
                    reprow_cveId = item["nvt"]["cve"]
                    # NEEDS FIXING CVE IS IMPORTED WITH AN EXTRA " IN THE START ----------------------------------------
                    reprow_cveId = '"' + reprow_cveId
                    if not db.session.query(
                            exists().where(CommonVulnerabilitiesAndExposures.CVEId == reprow_cveId)).scalar():
                        print("Continue2")
                        print(reprow_cveId)
                        continue
                    # my_report = db.session.query(VReport).filter_by(reportId=reprow_reportId).one()
                    my_cve = db.session.query(CommonVulnerabilitiesAndExposures).filter_by(CVEId=reprow_cveId).one()
                    if VulnerabilityReport.query.join(VulnerabilityReportVulnerabilitiesLink).join(
                            CommonVulnerabilitiesAndExposures).filter(
                        (VulnerabilityReportVulnerabilitiesLink.vreport_id == my_json_report.id) & (
                                VulnerabilityReportVulnerabilitiesLink.cve_id == my_cve.id)).first() is None:
                        my_link = VulnerabilityReportVulnerabilitiesLink(vreport_id=my_json_report.id, cve_id=my_cve.id)
                    else:
                        my_link = VulnerabilityReport.query.join(VulnerabilityReportVulnerabilitiesLink).join(
                            CommonVulnerabilitiesAndExposures).filter(
                            (VulnerabilityReportVulnerabilitiesLink.vreport_id == my_json_report.id) & (
                                    VulnerabilityReportVulnerabilitiesLink.cve_id == my_cve.id)).first()
                    my_link.VReport_assetID = item["host"]["asset"]["@asset_id"] if item["host"]["asset"][
                                                                                        "@asset_id"] is not None else ""
                    my_link.VReport_assetIp = obj["ip"] if obj["ip"] is not None else ""
                    my_link.VReport_port = item["port"] if item["port"] is not None else ""
                    my_link.comments = item["nvt"]["@oid"] if item["nvt"]["@oid"] is not None else ""
                    db.session.add(my_link)
                    print("Link Added")
                    print("my_link")
                    try:
                        db.session.commit()
                    except SQLAlchemyError as e:
                        db.session.rollback()
                        continue
                    # call API for CVE and CWE information
                    response = requests.get("https://services.nvd.nist.gov/rest/json/cve/1.0/" + item["nvt"]["cve"])
                    if response is not None and response.status_code == 200:
                        NVDreport = response.json()
                        # update CVE table with API values
                        impact = NVDreport['result']['CVE_Items'][0]['impact']
                        my_cve = db.session.query(CommonVulnerabilitiesAndExposures).filter_by(CVEId=reprow_cveId).one()
                        my_cve.severity = impact["baseMetricV2"]["severity"] if impact["baseMetricV2"][
                                                                                    "severity"] is not None else ""
                        my_cve.exploitabilityScore = impact["baseMetricV2"]['exploitabilityScore'] if \
                            impact["baseMetricV2"][
                                'exploitabilityScore'] is not None else ""
                        my_cve.impactScore = impact["baseMetricV2"]['impactScore'] if impact["baseMetricV2"][
                                                                                          'impactScore'] is not None else ""
                        my_cve.obtainAllPrivilege = impact["baseMetricV2"]['obtainAllPrivilege'] if \
                            impact["baseMetricV2"][
                                'obtainAllPrivilege'] is not None else ""
                        my_cve.obtainUserPrivilege = impact["baseMetricV2"]['obtainUserPrivilege'] if \
                            impact["baseMetricV2"][
                                'obtainUserPrivilege'] is not None else ""
                        my_cve.obtainOtherPrivilege = impact["baseMetricV2"]['obtainOtherPrivilege'] if \
                            impact["baseMetricV2"][
                                'obtainOtherPrivilege'] is not None else ""
                        my_cve.userInteractionRequired = impact["baseMetricV2"]['userInteractionRequired'] if \
                            impact["baseMetricV2"][
                                'userInteractionRequired'] is not None else ""
                        my_cve.accessVector = impact['baseMetricV2']['cvssV2']['accessVector'] if impact[
                                                                                                      'baseMetricV2'][
                                                                                                      'cvssV2'][
                                                                                                      'accessVector'] is not None else ""
                        my_cve.accessComplexity = impact['baseMetricV2']['cvssV2']['accessComplexity'] if impact[
                                                                                                              'baseMetricV2'][
                                                                                                              'cvssV2'][
                                                                                                              'accessComplexity'] is not None else ""
                        my_cve.authentication = impact['baseMetricV2']['cvssV2']['authentication'] if impact[
                                                                                                          'baseMetricV2'][
                                                                                                          'cvssV2'][
                                                                                                          'authentication'] is not None else ""
                        my_cve.confidentialityImpact = impact['baseMetricV2']['cvssV2']['confidentialityImpact'] if \
                            impact[
                                'baseMetricV2']['cvssV2']['confidentialityImpact'] is not None else ""
                        my_cve.integrityImpact = impact['baseMetricV2']['cvssV2']['integrityImpact'] if impact[
                                                                                                            'baseMetricV2'][
                                                                                                            'cvssV2'][
                                                                                                            'integrityImpact'] is not None else ""
                        my_cve.availabilityImpact = impact['baseMetricV2']['cvssV2']['availabilityImpact'] if impact[
                                                                                                                  'baseMetricV2'][
                                                                                                                  'cvssV2'][
                                                                                                                  'availabilityImpact'] is not None else ""
                        my_cve.baseScore = impact['baseMetricV2']['cvssV2']['baseScore'] if \
                            impact['baseMetricV2']['cvssV2']['baseScore'] is not None else ""
                        db.session.add(my_cve)
                        # Get CWEs and link them with CVE
                        for api_cve_desc, api_cveId, api_cweId in get_cwe_codes_from_API_report(NVDreport):
                            my_cve.description = api_cve_desc
                            cwe_number = api_cweId.split("CWE-", 1)[1].strip()
                            if cwe_number.isnumeric():
                                if db.session.query(
                                        exists().where(CommonWeaknessEnumeration.CWEId == cwe_number)).scalar():
                                    my_CWE_row = \
                                        db.session.query(CommonWeaknessEnumeration).filter_by(CWEId=cwe_number)[0]
                                else:
                                    my_CWE_row = CommonWeaknessEnumeration(CWEId=cwe_number)
                                    db.session.add(my_CWE_row)
                                if db.session.query(CommonVulnerabilitiesAndExposures).filter(
                                        VulnerabilitiesWeaknessLink.cwe_id == my_CWE_row.id,
                                        VulnerabilitiesWeaknessLink.cve_id == my_cve.id).first() is None:
                                    my_cVecWe = VulnerabilitiesWeaknessLink(cve_id=my_cve.id, cwe_id=my_CWE_row.id,
                                                                            date=datetime.utcnow())
                                else:
                                    my_cVecWe = db.session.query(CommonVulnerabilitiesAndExposures).filter(
                                        VulnerabilitiesWeaknessLink.cwe_id == my_CWE_row.id,
                                        VulnerabilitiesWeaknessLink.cve_id == my_cve.id).first()
                                    my_cVecWe.date = datetime.utcnow()
                                db.session.add(my_cVecWe)
                        db.session.commit()
            return 1


# endregion
# endregion

# region NVD API get_cwe_codes
def get_cwe_codes_from_API_report(APIreport):
    for item in APIreport['result']['CVE_Items']:
        itemdescr = item["cve"]["description"]["description_data"][0]["value"]
        for problem in item["cve"]['problemtype']['problemtype_data']:
            for descr in problem['description']:
                yield itemdescr, item["cve"]["CVE_data_meta"]["ID"], descr["value"]


# endregion

# region get_assets
# temporarily from VaasReport table
def get_assets():
    if db.session.query(VulnerabilityReportVulnerabilitiesLink).distinct(
            VulnerabilityReportVulnerabilitiesLink.VReport_assetID).count() > 0:
        list_of_assets = db.session.query(VulnerabilityReportVulnerabilitiesLink).distinct(
            VulnerabilityReportVulnerabilitiesLink.VReport_assetID)
        return list_of_assets
    else:
        return []
    # db.session.query(your_table.column1.distinct()).filter_by(column2 = 'some_column2_value').all()


#
# def get_assetsfromrepository():
#     if db.session.query().distinct(HardwareAsset.id).count() > 0:
#         list_of_assets = db.session.query(HardwareAsset).distinct(HardwareAsset.id)
#         return list_of_assets
#     else:
#         return -1


# region get Recommended CVEs for an asset
def get_cve_recommendations(asset_id):
    if db.session.query(VulnerabilityReportVulnerabilitiesLink).distinct(
            VulnerabilityReportVulnerabilitiesLink.cve_id).filter(
        VulnerabilityReportVulnerabilitiesLink.VReport_assetID == asset_id).count() > 0:
        list_of_cve = db.session.query(VulnerabilityReportVulnerabilitiesLink.cve_id).distinct(
            VulnerabilityReportVulnerabilitiesLink.cve_id).filter(
            VulnerabilityReportVulnerabilitiesLink.VReport_assetID == asset_id)
        result = db.session.query(CommonVulnerabilitiesAndExposures).filter(
            CommonVulnerabilitiesAndExposures.id.in_(list_of_cve))
        return result.all()
    else:
        return -1


# endregion

# region get Recommended CWEs for a selected CVE
def get_cwe_recommendations(selected_cve_id):
    if db.session.query(VulnerabilitiesWeaknessLink).distinct(VulnerabilitiesWeaknessLink.cwe_id).filter(
            VulnerabilitiesWeaknessLink.cve_id == selected_cve_id).count() > 0:
        list_of_cwe = db.session.query(VulnerabilitiesWeaknessLink.cwe_id).distinct(
            VulnerabilitiesWeaknessLink.cwe_id).filter(VulnerabilitiesWeaknessLink.cve_id == selected_cve_id)
        result = db.session.query(CommonWeaknessEnumeration).filter(CommonWeaknessEnumeration.id.in_(list_of_cwe))
        return result.all()
    else:
        return -1


# endregion

# region get Recommended CAPECs for a selected CVE
def get_capec_recommendations(selected_cve_id):
    for cwe_item in get_cwe_recommendations(selected_cve_id):
        capecList = []
        for capec_item in CommonAttackPatternEnumerationClassification.query.filter(
                CommonAttackPatternEnumerationClassification.relatedWeaknesses.like("%" + cwe_item.CWEId + "%")):
            capecList.append(capec_item.id)
    if capecList:
        result = db.session.query(CommonAttackPatternEnumerationClassification).distinct(
            CommonAttackPatternEnumerationClassification.capecId).filter(
            CommonAttackPatternEnumerationClassification.id.in_(capecList)) if db.session.query(
            CommonAttackPatternEnumerationClassification).filter(
            CommonAttackPatternEnumerationClassification.id.in_(capecList)) is not None else -1
        return result.all()
    else:
        return -1


# endregion

# region get Recommended CAPECs for a selected CVE

# One impact at a time saved
# scopes is an array
# NEEDS TO BE REACTIVATED, IT WORKS
# def save_capec_consequence(scopes, impact, notes):
#     scope_instances = []
#     for scope in scopes:
#         stored_scope = db.session.query(GiraScope).filter_by(name=scope).first()
#         if stored_scope is None:
#             new_scope = GiraScope(name=scope)
#             scope_instances.append(new_scope)
#             db.session.add(new_scope)
#         else:
#             scope_instances.append(stored_scope)
#
#     db.session.commit()
#     impact = impact[:-1]  # To be removed if bug that leaves ':' in the end of the asset is fixed
#     if db.session.query(GiraImpact).filter_by(name=impact).first() is None:
#         new_impact = GiraImpact(name=impact, note=notes)
#         for scope_instance in scope_instances:
#             new_impact.scopes.append(scope_instance)
#
#     db.session.commit()


# def get_capec_consequences():
#     capec_list = CommonAttackPatternEnumerationClassification.query.all()
#     for capec_entry in capec_list:
#         capec_consequence = capec_entry.consequences
#         if capec_consequence is None:
#             continue
#
#         temp_capec = capec_consequence[1:]
#         temp_capec = temp_capec[:-1]
#
#         temp_consequence_list = temp_capec.split(':SCOPE:')
#
#         # temp_scope contains scope that are linked to the next impact
#         temp_scope = []
#         for temp_it in temp_consequence_list:
#             if "TECHNICAL IMPACT:" in temp_it:
#                 temp_impact = temp_it
#                 temp_description = ""
#                 if "NOTE:" in temp_it:
#                     temp_impact_and_note = temp_it.split('NOTE:')
#                     temp_impact = temp_impact_and_note[0]
#                     temp_description = temp_impact_and_note[1]
#
#                 temp_scope_and_impact = temp_impact.split('TECHNICAL IMPACT:')
#                 temp_scope.append(temp_scope_and_impact[0])
#
#                 save_capec_consequence(temp_scope, temp_scope_and_impact[1], temp_description)
#                 print(temp_scope)
#                 print(temp_scope_and_impact[1])
#                 print(temp_description)
#                 print("-------------------")
#                 temp_scope = []
#
#             elif temp_it == "":
#                 continue
#             else:
#                 temp_scope.append(temp_it)


# def get_hardwareassets():
#     if db.session.query(HardwareAsset).distinct(
#             HardwareAsset.id).count() > 0:
#         list_of_hardwareassets = db.session.query(HardwareAsset).distinct(
#             HardwareAsset.id)
#         return list_of_hardwareassets
#     else:
#         return []


# endregion

# region Communication Functions
def sendDSSScore():
    asset = stix2.IPv4Address(
        # type="ipv4-addr",
        value="10.0.255.106"
    )
    attack = stix2.AttackPattern(
        # type="attack-pattern",
        name="Spear Phishing as Practiced by Adversary X",
        description="A particular form of spear phishing where the attacker claims that the target had won a contest, including personal details, to get them to click on a link.",
    )

    relationship = stix2.Relationship(
        # type="relationship",
        relationship_type="targets",
        source_ref=attack.id,
        target_ref=asset.id
    )

    scoring = {
        "score": "1",
        "impact": "high",
        "probability": "low"
    }
    rcra = stix2_custom.RCRAObjective(
        x_rcra_scoring=json.dumps(scoring)

    )

    bundle = stix2.Bundle(asset, attack, relationship, rcra)
    print(bundle, flush=True)
    stix2validator.validate_instance(bundle)
    SendKafkaReport(str(bundle), "rcra-report-topic")

    return 0


def sendDSSScoreTest():
    asset = stix2.IPv4Address(
        # type="ipv4-addr",
        value="10.0.255.106"
    )
    attack = stix2.AttackPattern(
        # type="attack-pattern",
        name="Spear Phishing as Practiced by Adversary X",
        description="A particular form of spear phishing where the attacker claims that the target had won a contest, including personal details, to get them to click on a link.",
    )

    relationship = stix2.Relationship(
        # type="relationship",
        relationship_type="targets",
        source_ref=attack.id,
        target_ref=asset.id
    )

    scoring = {
        "score": "20",
        "impact": "high",
        "probability": "low"
    }
    rcra = stix2_custom.RCRAObjective(
        x_rcra_scoring=json.dumps(scoring)

    )

    bundle = stix2.Bundle(asset, attack, relationship, rcra)
    print(bundle, flush=True)
    stix2validator.validate_instance(bundle)
    SendKafkaReport(str(bundle), "rcra-report-topic-test")

    return 0


def send_dss_alert():
    asset = stix2.IPv4Address(
        # type="ipv4-addr",
        value="10.0.255.106"
    )
    attack = stix2.AttackPattern(
        # type="attack-pattern",
        name="Spear Phishing as Practiced by Adversary X",
        description="A particular form of spear phishing where the attacker claims that the target had won a contest, including personal details, to get them to click on a link.",
    )

    relationship = stix2.Relationship(
        # type="relationship",
        relationship_type="targets",
        source_ref=attack.id,
        target_ref=asset.id
    )

    # scoring = {
    #     "score": "1",
    #     "impact": "high",
    #     "probability": "low"
    # }
    # rcra = stix2_custom.RCRAObjective(
    #     x_rcra_scoring=json.dumps(scoring)
    #
    # )

    bundle = stix2.Bundle(asset, attack, relationship)
    print(bundle, flush=True)
    stix2validator.validate_instance(bundle)
    # SendKafkaReport(str(bundle))
    return str(bundle)


def make_visualisation():
    """ Constructs example visualisation for Current Threats by impact level"""
    score = {
        "low_impact": "1",
        "medium_impact": "2",
        "high_impact": "2",
        "critical_impact": "1",
    }

    vis_1 = stix2_custom.RCRACurrentThreatsVis(
        x_rcra_threats=score
    )
    bundle = stix2.Bundle(vis_1)
    stix2validator.validate_instance(bundle)

    # print(bundle, flush=True)
    return str(bundle)


def make_visualisation_current_assets(assets):
    """ Constructs Visualisation for ID new-unverified asset alert """
    vis_1 = stix2_custom.RCRACurrentAssets(
        x_rcra_assets=assets
    )
    bundle = stix2.Bundle(vis_1)
    stix2validator.validate_instance(bundle)

    return str(bundle)


def send_asset_id_alert():
    """' Function Recieves New Detected Assets and Send new visualisation data to ID (all or uknown-unverified only?)"""
    return 1


def convert_database_items_to_json_table(items):
    """ Converts sqlalchemy entries to json ,
    Needs to be in an array to work"""
    if items:
        # print(items)
        columns = items[0].__table__.columns._data.keys()
        json_ready = []
        temp_json = {}
        # print(columns, flush=True)
        for item in items:
            # print(item)
            for column in columns:
                temp_json[column] = getattr(item, column)
            json_ready.append(temp_json.copy())
            # print(json_ready)

        return json_ready
    else:
        return []


def import_fixture_from_file(file_name):
    '''Function to imports Json from app/fixtures'''
    with open(os.path.join(os.getcwd(), "app", "fixtures", file_name + ".json"), encoding='utf-8') as json_file:
        return json.load(json_file)


def rcra_db_init():
    """Function is run in the _init_ file when server starts to initialise static table data"""
    print("Initiating Tamarin SSH Forward", flush=True)
    bash_com = './docker/expect.sh tamarin-prover ssh -4 -o "StrictHostKeyChecking no" -L 0.0.0.0:3005:localhost:3001 tamarin-prover@tamarin'
    cmd = ['./docker/expect.sh', 'tamarin-prover', 'ssh', '-4', '-o', "StrictHostKeyChecking no", '-L',
         '0.0.0.0:3005:localhost:3001', 'tamarin-prover@tamarin']
    process = subprocess.Popen(
        cmd, stdout=subprocess.PIPE)
    output, error = process.communicate()
    print(output)

    print("Initiating Database", flush=True)
    if RepoService.query.count() is not 0:
        print(RepoService.query.count())
        return "Already exists"

    # Adding Services
    to_add_services = import_fixture_from_file("repo_service")

    for service_json in to_add_services:
        print(service_json)
        to_add_service = RepoService(**service_json)
        db.session.add(to_add_service)

    # Adding Threats
    to_add_threats = import_fixture_from_file("repo_threat")

    for threat_json in to_add_threats:
        to_add_threat = RepoThreat(**threat_json)
        db.session.add(to_add_threat)

    # Adding Impacts
    to_add_impacts = import_fixture_from_file("repo_impact")

    for impact_json in to_add_impacts:
        to_add_impact = RepoImpact(**impact_json)
        db.session.add(to_add_impact)

    # Adding Objectives
    to_add_objectives = import_fixture_from_file("repo_objective")

    for objective_json in to_add_objectives:
        to_add_objective = RepoObjective(**objective_json)
        db.session.add(to_add_objective)

    to_add_objective_options = import_fixture_from_file("repo_objectives_option")

    for objectives_option_json in to_add_objective_options:
        to_add_objectives_option = RepoObjectivesOptions(**objectives_option_json)
        db.session.add(to_add_objectives_option)

    db.session.commit()


def start_risk_assessment(threat_id, asset_id):
    diag = gum.InfluenceDiagram()
    try:
        this_risk_assessment = RepoRiskAssessment.query.filter_by(repo_threat_id=threat_id,
                                                                  repo_asset_id=asset_id).first()
    except SQLAlchemyError:
        return "SQLAlchemyError"
    ## Node creation
    this_asset = this_risk_assessment.asset
    this_threat = this_risk_assessment.threat

    # Node creation threat exposure
    exposureNodeId = "te" + str(this_threat.id)
    diag.add(gum.LabelizedVariable(exposureNodeId, this_threat.name, 2))

    # Node creation responses
    try:
        these_responses = RepoResponse.query.filter_by(threat_id=threat_id).all()
    except SQLAlchemyError:
        return "SQLAlchemyError"

    nodeId = "re"
    diag.addDecisionNode(gum.LabelizedVariable(nodeId, these_responses[1].name, 2))

    # for response in these_responses:
    #     nodeId = "re" + str(response.id)
    #     diag.addDecisionNode(gum.LabelizedVariable(nodeId, response.name, 2))

    # Node creation materialisations
    try:
        these_materialisations = RepoMaterialisation.query.filter_by(threat_id=threat_id).all()
    except SQLAlchemyError:
        return "SQLAlchemyError"

    for materialisation in these_materialisations:
        nodeId = "mat" + str(materialisation.id)
        diag.add(gum.LabelizedVariable(nodeId, materialisation.name, 2))

    # Node creation consequences
    try:
        these_consequences = RepoConsequence.query.filter_by(threat_id=threat_id).all()
    except SQLAlchemyError:
        return "SQLAlchemyError"

    for consequence in these_consequences:
        nodeId = "con" + str(consequence.id)
        diag.add(gum.LabelizedVariable(nodeId, consequence.name, 2))

    # Node creation assets
    try:
        these_services = RepoService.query.filter(RepoService.assets.any(id=asset_id)).all()
    except SQLAlchemyError:
        return "SQLAlchemyError"

    for service in these_services:
        nodeId = "serv" + str(service.id)
        diag.addDecisionNode(gum.LabelizedVariable(nodeId, service.name, 2))

    # Node creation impacts
    try:
        these_impacts = RepoImpact.query.all()
    except SQLAlchemyError:
        return "SQLAlchemyError"

    for impact in these_impacts:
        nodeId = "imp" + str(impact.id)
        diag.add(gum.LabelizedVariable(nodeId, impact.name, 3))

    # Node creation objectives
    try:
        these_objectives = RepoObjective.query.all()
    except SQLAlchemyError:
        return "SQLAlchemyError"

    for objective in these_objectives:
        nodeId = "obj" + str(objective.id)
        diag.add(gum.LabelizedVariable(nodeId, objective.name, 3))

    ##Node Linking
    # Link Exposure and response to materialisation
    for materialisation in these_materialisations:
        nodeId = "mat" + str(materialisation.id)
        diag.addArc(exposureNodeId, nodeId)

    # for response in these_responses:
    for materialisation in these_materialisations:
        nodeId = "re"
        nodeMatId = "mat" + str(materialisation.id)
        diag.addArc(nodeId, nodeMatId)

    # Link Mat and Re to Cons
    for consequence in these_consequences:
        nodeConsId = "con" + str(consequence.id)
        nodeMatId = "mat" + str(consequence.materialisation_id)
        nodeReId = "re"

        diag.addArc(nodeReId, nodeConsId)
        diag.addArc(nodeMatId, nodeConsId)

    # Link cons and service in impacts
    for service in these_services:
        nodeServId = "serv" + str(service.id)
        try:
            these_related_impacts = RepoImpact.query.filter(RepoImpact.services.any(id=service.id)).all()
        except SQLAlchemyError:
            return "SQLAlchemyError"

        for impact in these_related_impacts:
            nodeImpactId = "imp" + str(impact.id)
            diag.addArc(nodeServId, nodeImpactId)

    for consequence in these_consequences:
        nodeConsId = "con" + str(consequence.id)
        try:
            these_related_impacts = RepoImpact.query.filter(RepoImpact.consequences.any(id=consequence.id)).all()
        except SQLAlchemyError:
            return "SQLAlchemyError"

        for impact in these_related_impacts:
            nodeImpactId = "imp" + str(impact.id)
            diag.addArc(nodeConsId, nodeImpactId)

    # Link objective to imp
    for impact in these_impacts:
        nodeImpactId = "imp" + str(impact.id)
        try:
            these_related_objectives = RepoObjective.query.filter(RepoObjective.impacts.any(id=impact.id))
        except SQLAlchemyError:
            return "SQLAlchemyError"

        for objective in these_related_objectives:
            nodeObjectiveId = "obj" + str(objective.id)
            diag.addArc(nodeImpactId, nodeObjectiveId)

    ## Node Value Filling
    # Exposure Node Values
    diag.cpt("te1").fillWith([0.7, 0.5])

    # Materialisation Node Values
    for materialisation in these_materialisations:
        # print("----- Matinfo ------")
        # print(nodeImpactId)
        # print(nodeObjectiveId)
        nodeMatId = "mat" + str(materialisation.id)
        nodeReId = "re"
        try:
            these_materialisation_values = RepoRiskThreatAssetMaterialisation.query.filter_by(repo_asset_id=asset_id,
                                                                                              repo_threat_id=threat_id,
                                                                                              repo_materialisation_id=materialisation.id).all()
        except SQLAlchemyError:
            return "SQLAlchemyError"
        for node_value in these_materialisation_values:
            if node_value.threat_occurrence is True:
                occurance_bool_num = 1
            else:
                occurance_bool_num = 0

            # response shouldnt work like that this needs a bit of a rework
            if node_value.repo_response_id % 2 == 0:
                response_bool_num = 1
            else:
                response_bool_num = 0

            diag.cpt(nodeMatId)[{exposureNodeId: occurance_bool_num, nodeReId: response_bool_num}] = [node_value.prob,
                                                                                                      1 - node_value.prob]

        print(these_materialisation_values)

    # Consequence Node Values
    for consequence in these_consequences:
        # print("----- Matinfo ------")
        # print(nodeImpactId)
        # print(nodeObjectiveId)
        nodeConsId = "con" + str(consequence.id)
        nodeReId = "re"
        try:
            these_cosnequence_values = RepoRiskThreatAssetConsequence.query.filter_by(repo_asset_id=asset_id,
                                                                                      repo_threat_id=threat_id,
                                                                                      repo_consequence_id=consequence.id).all()
        except SQLAlchemyError:
            return "SQLAlchemyError"

        for node_value in these_cosnequence_values:
            if node_value.threat_occurrence is True:
                occurance_bool_num = 1
            else:
                occurance_bool_num = 0

            # response shouldnt work like that this needs a bit of a rework
            if node_value.repo_response_id % 2 == 0:
                response_bool_num = 1
            else:
                response_bool_num = 0

            nodeMatId = "mat" + str(node_value.repo_consequence.materialisation_id)
            diag.cpt(nodeConsId)[{nodeMatId: occurance_bool_num, nodeReId: response_bool_num}] = [node_value.prob,
                                                                                                  1 - node_value.prob]
        print(these_cosnequence_values)

    # Impact Node Values
    for impact in these_impacts:
        nodeImpactId = "imp" + str(impact.id)

        array_impact_calculation = []

        print("Related services are")
        print(these_services)

        print("Related Consequence are")
        print(these_consequences)

        for repo_temp_service in these_services:
            if not array_impact_calculation:
                temp_to_add_1 = {"service": repo_temp_service, "state": True}
                temp_to_add_2 = {"service": repo_temp_service, "state": False}
                array_impact_calculation.append([temp_to_add_1])
                array_impact_calculation.append([temp_to_add_2])
            else:
                temp_impact_array = deepcopy(array_impact_calculation)
                for to_be_added in temp_impact_array:
                    to_be_added.append({"service": repo_temp_service, "state": True})

                for to_be_added in array_impact_calculation:
                    to_be_added.append({"service": repo_temp_service, "state": False})

                array_impact_calculation = array_impact_calculation + temp_impact_array

        for repo_temp_consequence in these_consequences:
            if not array_impact_calculation:
                temp_to_add_1 = {"consequence": repo_temp_consequence, "state": True}
                temp_to_add_2 = {"consequence": repo_temp_consequence, "state": False}
                array_impact_calculation.append([temp_to_add_1])
                array_impact_calculation.append([temp_to_add_2])
            else:
                temp_impact_array = deepcopy(array_impact_calculation)
                for to_be_added in temp_impact_array:
                    to_be_added.append({"consequence": repo_temp_consequence, "state": True})

                for to_be_added in array_impact_calculation:
                    to_be_added.append({"consequence": repo_temp_consequence, "state": False})

                array_impact_calculation = array_impact_calculation + temp_impact_array

        print("--- FINAL ARRAY ---")
        for temp in array_impact_calculation:
            print(temp)

        joined = db.session.query(RepoAssetThreatConsequenceServiceImpactRelationship,
                                  RepoAssetThreatConsequenceServiceImpactRelationshipConsequenceManyToMany,
                                  RepoAssetThreatConsequenceServiceImpactRelationshipServiceManyToMany).join(
            RepoAssetThreatConsequenceServiceImpactRelationshipConsequenceManyToMany,
            RepoAssetThreatConsequenceServiceImpactRelationshipServiceManyToMany).filter(
            RepoAssetThreatConsequenceServiceImpactRelationship.repo_threat_id == threat_id,
            RepoAssetThreatConsequenceServiceImpactRelationship.repo_impact_id == impact.id,
            RepoAssetThreatConsequenceServiceImpactRelationship.repo_asset_id == asset_id,
        ).all()

        concatted = {}

        for temp_joined in joined:
            # print("Single Line")
            # print("Inner Line")
            if temp_joined[0] not in concatted:
                concatted[temp_joined[0]] = []
            for inner_joined in temp_joined:
                if inner_joined is temp_joined[0]:
                    continue
                # print(concatted[temp_joined[0]])
                if type(inner_joined) is RepoAssetThreatConsequenceServiceImpactRelationshipConsequenceManyToMany:
                    # inner_joined_arrayed = ['cons', inner_joined.repo_consequence_id, inner_joined.repo_consequence_state]
                    inner_joined_arrayed = {"consequence": {'id': inner_joined.repo_consequence_id,
                                                            'name': inner_joined.repo_consequence.name,
                                                            'threat_id': inner_joined.repo_consequence.threat_id,
                                                            'materialisation_id': inner_joined.repo_consequence.materialisation_id
                                                            },
                                            "state": inner_joined.repo_consequence_state}
                else:
                    # inner_joined_arrayed = ['serv', inner_joined.repo_service_id, inner_joined.repo_service_state]
                    inner_joined_arrayed = {
                        "service": {'id': inner_joined.repo_service_id, 'name': inner_joined.repo_service.name},
                        "state": inner_joined.repo_service_state}
                if inner_joined_arrayed not in concatted[temp_joined[0]]:
                    concatted[temp_joined[0]].append(inner_joined_arrayed)
        # print("------------ RESULTS ARE ----------")
        # print(concatted.items())

        for concatted_entry_key, concatted_entry_value in concatted.items():
            # print("------Comparison------")
            # print(concatted_entry_key)
            # print(concatted_entry_value)
            impact_node_value = []
            impact_node_id = {}

            for temp_entry in concatted_entry_value:
                # state_int = 0
                if temp_entry['state'] is False:
                    state_int = 0
                else:
                    state_int = 1

                if 'consequence' in temp_entry:
                    nodeTempImpactId = "con" + str(temp_entry['consequence']['id'])
                else:
                    nodeTempImpactId = "serv" + str(temp_entry['service']['id'])

                impact_node_id[nodeTempImpactId] = state_int

            # print()
            impact_node_value.append(concatted_entry_key.low_prob)
            # objective_node_value.append(1 - concatted_entry_key.low_prob)
            impact_node_value.append(concatted_entry_key.med_prob)
            # objective_node_value.append(1 - concatted_entry_key.med_prob)
            impact_node_value.append(concatted_entry_key.high_prob)
            # objective_node_value.append(1 - concatted_entry_key.high_prob)

            print("---------- TO ADD ERROR ----------------")
            print(impact_node_id)
            print(impact_node_value)
            print(nodeImpactId)
            diag.cpt(nodeImpactId)[impact_node_id] = impact_node_value

    # Objective  Node Values
    for objective in these_objectives:
        nodeObjectiveId = "obj" + str(objective.id)

        joined = db.session.query(RepoObjectiveImpactRelationship,
                                  RepoObjectiveImpactRelationshipImpactManyToMany) \
            .join(RepoObjectiveImpactRelationshipImpactManyToMany) \
            .filter(
            RepoObjectiveImpactRelationship.repo_objective_id == objective.id,
        ).all()

        concatted = {}

        for temp_joined in joined:
            # print("Single Line")
            # print("Inner Line")
            if temp_joined[0] not in concatted:
                concatted[temp_joined[0]] = []
            for inner_joined in temp_joined:
                if inner_joined is temp_joined[0]:
                    continue
                # print(concatted[temp_joined[0]])
                if inner_joined.repo_impact_state == 0:
                    temp_state = "low"
                elif inner_joined.repo_impact_state == 1:
                    temp_state = "med"
                else:
                    temp_state = "high"
                inner_joined_arrayed = {"impact": inner_joined.repo_impact,
                                        "state": temp_state}

                if inner_joined_arrayed not in concatted[temp_joined[0]]:
                    concatted[temp_joined[0]].append(inner_joined_arrayed)

        for concatted_entry_key, concatted_entry_value in concatted.items():
            # print("------Comparison------")
            # print(concatted_entry_key)
            # print(concatted_entry_value)
            objective_node_value = []
            objective_node_id = {}

            for temp_entry in concatted_entry_value:
                # state_int = 0
                if temp_entry['state'] == 'low':
                    state_int = 0
                if temp_entry['state'] == 'med':
                    state_int = 1
                if temp_entry['state'] == 'high':
                    state_int = 2

                nodeImpactId = "imp" + str(temp_entry['impact'].id)

                objective_node_id[nodeImpactId] = state_int

            # print()
            objective_node_value.append(concatted_entry_key.low_prob)
            # objective_node_value.append(1 - concatted_entry_key.low_prob)
            objective_node_value.append(concatted_entry_key.med_prob)
            # objective_node_value.append(1 - concatted_entry_key.med_prob)
            objective_node_value.append(concatted_entry_key.high_prob)
            # objective_node_value.append(1 - concatted_entry_key.high_prob)

            diag.cpt(nodeObjectiveId)[objective_node_id] = objective_node_value

    # Print Diagram
    diag.saveBIFXML(os.path.join("out", "GiraDynamic.bifxml"))

    ie = gum.ShaferShenoyLIMIDInference(diag)

    no_forgetting_array = []

    no_forgetting_array.append("re")

    for service in these_services:
        nodeServId = "serv" + str(service.id)
        no_forgetting_array.append(nodeServId)

    # for response in these_responses:

    ie.addNoForgettingAssumption(no_forgetting_array)

    print("Is this solvable =" + str(ie.isSolvable()))
    ie.addEvidence('te1', 1)
    ie.addEvidence('re', 0)

    ie.makeInference()

    print("-------- INFERENCE RESULTS ----------")
    print(ie.posterior('obj1'))
    print(ie.posterior('obj2'))
    print(ie.posterior('obj3'))
    print(ie.posterior('obj4'))
    print(ie.posterior('obj5'))
    # Print Graph
    # with open(os.path.join("out", "GiraDynamic.bifxml"), "r") as out:
    # print(out.read())
    # try:
    #     mat_nodes = RepoRiskThreatAssetMaterialisation.query.filter_by(repo_asset_id=asset_id,
    #                                                                                       repo_threat_id=threat_id,
    #                                                                                       repo_materialisation_id=
    #                                                                                       deconstructedId[1],
    #                                                                                       repo_response_id=
    #                                                                                       deconstructedId[2],
    #                                                                                       threat_occurrence=to_add_threat_occurence_bool).first()
    # except SQLAlchemyError:
    #     return "SQLAlchemyError"
    #

# def repo_check_dtm_asset_exits_and_add(dtm_object, assets):
#     """This function uses a single object of the DTM and the retrieved database assets
#     as input and will check if the described asset exists in the database"""
#     if dtm_object in assets:


# url = "http://sphinx-kubernetes.intracom-telecom.com:8080/SMPlatform/manager/rst/Authentication"
# payload = {
#     'username': 'testR1',
#     'password': 'testR1123!@'
# }
# response = requests.request("POST", url, data=payload)
# selectedticket = response.json()
# requestedTicket = selectedticket["data"]
#
# print("---------------------------------------", flush=True)
# print("Login ticket is: ", requestedTicket, flush=True)
# print("---------------------------------------", flush=True)
#
# # Need endpoint of dss
# url = "http://sphinx-dss-service:5000/-"
# params = {
#     'requestedservice': 'DSS',
#     'requestedTicket': requestedTicket
# }
#
# data = jsonify({'alert-1': [
#     {"asset": "server 1", "threat-level": "high", "date-time": "-", "type": "--"}
# ],
# })
# response = requests.request("POST", url, params=params, data=data)
#
# if response.status_code != 200:
#     return 1
# else:
#     return 0

# endregion


# region Test area
# db.create_all()
# x= v_report("Json_texts/report1.json")# for x in v_report("Json_texts/report1.json"):
# print(x)

# for y in get_assetsfromrepository():
#     print(y.id)

# for y in get_cve_recommendations('f080c7b3-3038-4a52-8b14-4397136c9dad'):
#     print(y.CVEId, y.id)
#
# for y in get_cwe_recommendations('170528'):
#     print(y.CWEId, y.id)

# for xx in CAPEC.query.filter(CAPEC.relatedWeaknesses.like("%200%")):
#     print(xx.capecId, xx.relatedWeaknesses)
# i = 0
# for xx in get_capec_recommendations('170528'):
#     i = i + 1
#     print(i, xx.capecId, xx.relatedWeaknesses)

# for y in db.session.query(VReportCVELink).distinct(VReportCVELink.VReport_assetID):
#     print(y.VReport_assetID, y.VReport_port, y.VReport_assetIp)

#
# for y in cVecWe.query.all():
#     print(y.cwe_id, y.cve_id, y.date)
#
# x= db.session.query(CVE).filter_by(CVEId="CVE-1999-0524").one()
# print(x.id, x.CVEId, x.description)
# for y in CVE.query.all():
#     print(y.CVEId, y.description)

# db.create_all()
# x = CAPEC_excel_insertData('xlsx_texts/CAPEC-Domains of Attack-3000.xlsx')
# print('Return: {}'.format(x))
# for xx in CommonAttackPatternEnumerationClassification.query.all():
#     print(xx.id, xx.capecId, xx.name, xx.relatedWeaknesses)
#
# x = CWE_excel_insertData('xlsx_texts/CWE-Research Concepts-1000.xlsx')
# print('Return: {}'.format(x))
# for xx in CommonWeaknessEnumeration.query.all():
#     print(xx.id, xx.CWEId, xx.name)
#
# x = CVE_excel_insertData('xlsx_texts/CVE-allitems.xlsx')
# print('Return: {}'.format(x))
# for xx in CommonVulnerabilitiesAndExposures.query.all():
#     print(xx.id, xx.CVEId, xx.status)

# endregion
