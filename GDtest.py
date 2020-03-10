import json
import openpyxl
import requests


def v_report(fpath):
    with open(fpath, "r") as fp:
        obj = json.load(fp)
        for item in obj['report']['report']['results']['result']:
            yield obj["ip"], item["host"]["asset"]["@asset_id"], item["nvt"]["cvss_base"], item["nvt"]["cve"], \
                  item["nvt"]["@oid"], item["name"]
            # print(item["host"]["asset"]["@asset_id"],
            #       ' - ' + item["nvt"]["cvss_base"] + ' - ' + item["nvt"]["cve"] + ' - ' + item["nvt"]["@oid"],
            #       ' - ' + item["name"])


def my_excel_read(xpath, xcol, cwe_code):
    theFile = openpyxl.load_workbook(xpath)
    print(theFile.sheetnames)
    print("All sheet names {} ".format(theFile.sheetnames))
    currentSheet = theFile.active
    # for rowidx in range(1, currentSheet.max_row + 1):
    # ws['B'] will return all cells on the B column until the last one (similar to max_row but it's only for the B column)
    for cell in currentSheet[xcol]:
        if cell.value is not None:  # We need to check that the cell is not empty.
            if cwe_code in cell.value:  # Check if the value of the cell contains the text 'Table'
                # print('Found header with name: {} at row: {} and column: {}. In cell {}'.format(cell.value, cell.row, cell.column, cell))
                yield currentSheet.cell(cell.row, 1).value, currentSheet.cell(cell.row, 2).value


def get_cwe_codes(myreport):
    for item in myreport['result']['CVE_Items']:
        itemdescr = item["cve"]["description"]["description_data"][0]["value"]
        for problem in item["cve"]['problemtype']['problemtype_data']:
            for descr in problem['description']:
                yield itemdescr, item["cve"]["CVE_data_meta"]["ID"], descr["value"]


# print(response.json())


###### console tests
# for VAaaS report - get CVE codes
# for ip, asset, cvss, cve, oid, name in v_report("app/Json_texts/report1.json"):
##     print(ip, asset, cvss, cve, oid, name)
##### through NVD API - get CWE codes
response = requests.get("https://services.nvd.nist.gov/rest/json/cve/1.0/CVE-2009-3421")
if response.status_code == 200:
    myreport = response.json()
    for cve, id, cwe in get_cwe_codes(myreport):
        cwe_number = cwe.split("CWE-", 1)[1].strip()
        if cwe_number.isnumeric():
            print(cve, id, cwe.split("CWE-", 1)[1].strip())
######## for CWE - get CAPEC patterns
            for x, y in my_excel_read('app/xlsx_texts/CAPEC-Domains of Attack-3000.xlsx', 'R', cwe_number):
                print(x, y)

